"""요건추적표(RTM) 엑셀 렌더러.

순수 함수: (검증된 RTMDocument, 템플릿 경로) → 출력 .xlsx 파일.
템플릿의 서식 기준 행을 복제해 값만 주입하며, 서식을 코드로 새로 그리지 않는다.
"""

from copy import copy
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet

from app.exceptions import RenderError
from app.schemas.rtm import RTMDocument, RTMRow

# 템플릿 구조 상수 (backend/templates/rtm.xlsx 와 일치해야 함)
STYLE_ROW = 10  # 서식 기준 행 — 이 행의 서식을 데이터 행 전체에 복제한다
NUM_COLUMNS = 9
SHEET_NAME = "요건추적표"
STAGE_MARK = "O"  # 단계 반영 여부 표기 (반영 시)

# 표지 정보 영역의 값 셀 위치 (test_scenario 템플릿과 동일한 배치)
COVER_CELLS = {
    "project_name": "B5",
    "system_name": "G5",
    "author": "B6",
    "written_date": "G6",
}

# 셀 서식 복제 대상 속성 (openpyxl StyleProxy 는 copy 후 대입해야 한다)
_STYLE_ATTRS = ("font", "border", "fill", "alignment", "number_format", "protection")


def render_rtm(doc: RTMDocument, template_path: Path, output_path: Path) -> Path:
    """템플릿에 표지 정보와 추적 행을 주입해 출력 파일을 생성한다."""
    if not template_path.is_file():
        raise RenderError(f"템플릿 파일이 없습니다: {template_path}")

    wb = load_workbook(template_path)
    if SHEET_NAME not in wb.sheetnames:
        raise RenderError(f"템플릿에 '{SHEET_NAME}' 시트가 없습니다: {template_path}")
    ws = wb[SHEET_NAME]
    _fill_cover(ws, doc)
    _fill_rows(ws, doc.rows)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path


def _fill_cover(ws: Worksheet, doc: RTMDocument) -> None:
    """표지 정보 영역(프로젝트명/시스템명/작성자/작성일)에 값을 주입한다."""
    ws[COVER_CELLS["project_name"]] = doc.project_name
    ws[COVER_CELLS["system_name"]] = doc.system_name
    ws[COVER_CELLS["author"]] = doc.author
    ws[COVER_CELLS["written_date"]] = doc.written_date.isoformat()


def _fill_rows(ws: Worksheet, rows: list[RTMRow]) -> None:
    """서식 기준 행의 서식을 복제하면서 추적 행을 행 단위로 주입한다."""
    styles = [
        _capture_style(ws.cell(row=STYLE_ROW, column=col)) for col in range(1, NUM_COLUMNS + 1)
    ]
    row_height = ws.row_dimensions[STYLE_ROW].height

    for offset, rtm_row in enumerate(rows):
        row = STYLE_ROW + offset
        for col, value in enumerate(_row_to_cells(rtm_row), start=1):
            cell = ws.cell(row=row, column=col, value=value)
            _apply_style(cell, styles[col - 1])
        ws.row_dimensions[row].height = row_height


def _row_to_cells(rtm_row: RTMRow) -> list[str]:
    """추적 행 1건을 템플릿 열 순서의 값 목록으로 변환한다."""
    stage = rtm_row.stage_reflection
    return [
        rtm_row.req_id,
        rtm_row.req_name,
        "\n".join(rtm_row.screen_ids),
        "\n".join(rtm_row.program_ids),
        "\n".join(rtm_row.tc_ids),
        STAGE_MARK if stage.analysis else "",
        STAGE_MARK if stage.design else "",
        STAGE_MARK if stage.implementation else "",
        STAGE_MARK if stage.test else "",
    ]


def _capture_style(cell: Cell) -> dict[str, Any]:
    # Any 사용 사유: Font/Border 등 서로 다른 openpyxl 스타일 타입을 속성명으로 묶어 다룬다
    return {attr: copy(getattr(cell, attr)) for attr in _STYLE_ATTRS}


def _apply_style(cell: Cell, style: dict[str, Any]) -> None:
    for attr, value in style.items():
        setattr(cell, attr, value)
