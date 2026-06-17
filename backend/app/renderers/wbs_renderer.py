"""WBS(작업분해구조) 엑셀 렌더러.

순수 함수: (검증된 WBSDocument, 템플릿 경로) → 출력 .xlsx 파일.
템플릿의 서식 기준 행을 복제해 값만 주입한다.

**계층 번호(1.1.2)·일정(시작/종료일)·요약 태스크 공수 합산을 렌더러가 계산한다** (CLAUDE.md 원칙).
- 계층 번호: 트리 전위 순회로 1, 1.1, 1.1.2 … 부여.
- 일정: 작업(leaf) 태스크는 선행 종료 다음 날(없으면 프로젝트 시작일)부터 기간만큼.
  요약 태스크는 자식 일정의 최소 시작 ~ 최대 종료로 산정. (PoC 는 달력일 기준, 휴일 미고려)
- 공수: 작업 태스크는 입력값, 요약 태스크는 자손 작업 합산.
- 선행 태스크 열: 선행 id 를 계산된 계층 번호로 표시한다.
"""

from copy import copy
from datetime import date, timedelta
from pathlib import Path
from typing import Any

import holidays
from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet

from app.exceptions import RenderError
from app.schemas.wbs import WBSDocument, WBSTask

_KR_HOLIDAYS = holidays.country_holidays("KR")


def is_workday(d: date) -> bool:
    """주말과 대한민국 공휴일(대체공휴일 포함)이 아니면 True."""
    if d.weekday() >= 5:
        return False
    if d in _KR_HOLIDAYS:
        return False
    return True


def next_workday(d: date) -> date:
    """주어진 날짜 다음 날부터 시작하여 가장 처음 도래하는 영업일을 반환한다."""
    curr = d + timedelta(days=1)
    while not is_workday(curr):
        curr += timedelta(days=1)
    return curr


def get_workday_start(d: date) -> date:
    """주어진 날짜가 영업일이면 그대로, 아니면 그 이후 최초 영업일을 반환한다."""
    curr = d
    while not is_workday(curr):
        curr += timedelta(days=1)
    return curr


def add_workdays(start: date, days: int) -> date:
    """start 일자를 포함하여 days 영업일만큼 일정을 진행시켰을 때의 종료일을 계산한다.

    days >= 1 일 때 작동한다.
    """
    if days <= 1:
        return start
    curr = start
    workdays_added = 1
    while workdays_added < days:
        curr += timedelta(days=1)
        if is_workday(curr):
            workdays_added += 1
    return curr


# 템플릿 구조 상수 (backend/templates/wbs.xlsx 와 일치해야 함)
STYLE_ROW = 9
NUM_COLUMNS = 8
SHEET_NAME = "WBS"

COVER_CELLS = {
    "project_name": "B5",
    "system_name": "F5",
    "author": "B6",
    "written_date": "F6",
}

_STYLE_ATTRS = ("font", "border", "fill", "alignment", "number_format", "protection")


class _Computed:
    """렌더러가 계산한 태스크별 값(계층번호·일정·공수)."""

    def __init__(self) -> None:
        self.number: dict[str, str] = {}  # id → 계층 번호 ("1.1.2")
        self.start: dict[str, date] = {}
        self.end: dict[str, date] = {}
        self.effort: dict[str, float] = {}
        self.order: list[WBSTask] = []  # 전위 순회 순서 (출력 행 순서)


def render_wbs(doc: WBSDocument, template_path: Path, output_path: Path) -> Path:
    """템플릿에 표지 정보와 (계산된) WBS 행을 주입해 출력 파일을 생성한다."""
    if not template_path.is_file():
        raise RenderError(f"템플릿 파일이 없습니다: {template_path}")

    wb = load_workbook(template_path)
    if SHEET_NAME not in wb.sheetnames:
        raise RenderError(f"템플릿에 '{SHEET_NAME}' 시트가 없습니다: {template_path}")
    ws = wb[SHEET_NAME]

    computed = _compute(doc)
    _fill_cover(ws, doc)
    _fill_rows(ws, computed)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path


def _compute(doc: WBSDocument) -> _Computed:
    """계층 번호·일정·공수를 계산한다 (결정론적, LLM 미사용)."""
    c = _Computed()
    _assign_numbers(doc.tasks, "", c)
    _schedule_leaves(doc, c)
    for task in doc.tasks:
        _rollup(task, c)
    return c


def _assign_numbers(tasks: list[WBSTask], prefix: str, c: _Computed) -> None:
    """전위 순회로 계층 번호를 부여하고 출력 순서를 기록한다."""
    for i, task in enumerate(tasks, start=1):
        number = f"{prefix}{i}"
        c.number[task.id] = number
        c.order.append(task)
        if task.children:
            _assign_numbers(task.children, f"{number}.", c)


def _schedule_leaves(doc: WBSDocument, c: _Computed) -> None:
    """작업(leaf) 태스크의 시작/종료일을 선행 의존을 반영해 전진 계산한다.

    영업일 기준(주말 및 대한민국 법정 공휴일 배제)으로 시작일과 종료일을 산정한다.
    """
    leaves = [t for t in c.order if not t.is_summary]
    tasks_dict = {t.id: t for t in c.order}
    parent_map = {}
    for t in c.order:
        for child in t.children:
            parent_map[child.id] = t.id

    def get_ancestors(tid: str) -> list[str]:
        ancestors = [tid]
        curr = tid
        while curr in parent_map:
            curr = parent_map[curr]
            ancestors.append(curr)
        return ancestors

    def get_leaf_descendants(tid: str) -> set[str]:
        t = tasks_dict[tid]
        if not t.is_summary:
            return {t.id}
        res = set()
        for child in t.children:
            res.update(get_leaf_descendants(child.id))
        return res

    # Resolve predecessors to leaf task IDs
    resolved_preds: dict[str, set[str]] = {}
    for task in leaves:
        preds = set()
        for ancestor_id in get_ancestors(task.id):
            ancestor_task = tasks_dict[ancestor_id]
            for pred in ancestor_task.predecessors:
                if pred in tasks_dict:
                    preds.update(get_leaf_descendants(pred))
        resolved_preds[task.id] = preds

    pending = list(leaves)
    while pending:
        progressed = False
        for task in list(pending):
            preds = resolved_preds[task.id]
            if all(p in c.end for p in preds):
                # 프로젝트 시작일 자체가 공휴일/주말일 수 있으므로 최초 영업일로 보정
                start = get_workday_start(doc.start_date)
                for pred in preds:
                    # 선행 작업 종료일 다음 날의 최초 영업일을 계산
                    candidate = next_workday(c.end[pred])
                    if candidate > start:
                        start = candidate
                c.start[task.id] = start
                # 시작일을 포함한 duration_days 영업일 후의 날짜 계산
                c.end[task.id] = add_workdays(start, task.duration_days)
                c.effort[task.id] = task.effort_md
                pending.remove(task)
                progressed = True
        if not progressed:  # 스키마 검증으로 순환은 없어야 하지만 방어적으로 중단
            raise RenderError("선행 의존을 해소할 수 없습니다 (순환 가능성)")


def _rollup(task: WBSTask, c: _Computed) -> tuple[date, date, float]:
    """요약 태스크의 일정/공수를 자손에서 합산한다 (후위 순회)."""
    if not task.is_summary:
        return c.start[task.id], c.end[task.id], c.effort[task.id]
    spans = [_rollup(child, c) for child in task.children]
    start = min(s for s, _, _ in spans)
    end = max(e for _, e, _ in spans)
    effort = sum(eff for _, _, eff in spans)
    c.start[task.id] = start
    c.end[task.id] = end
    c.effort[task.id] = effort
    return start, end, effort


def _fmt_effort(value: float) -> str:
    """공수를 정수면 정수로, 아니면 소수 1자리로 표기한다."""
    return str(int(value)) if value == int(value) else f"{value:.1f}"


def _fill_cover(ws: Worksheet, doc: WBSDocument) -> None:
    ws[COVER_CELLS["project_name"]] = doc.project_name
    ws[COVER_CELLS["system_name"]] = doc.system_name
    ws[COVER_CELLS["author"]] = doc.author
    ws[COVER_CELLS["written_date"]] = doc.written_date.isoformat()


def _fill_rows(ws: Worksheet, c: _Computed) -> None:
    """서식 기준 행의 서식을 복제하면서 WBS 행을 주입한다."""
    styles = [
        _capture_style(ws.cell(row=STYLE_ROW, column=col)) for col in range(1, NUM_COLUMNS + 1)
    ]
    row_height = ws.row_dimensions[STYLE_ROW].height

    for offset, task in enumerate(c.order):
        row = STYLE_ROW + offset
        for col, value in enumerate(_row_to_cells(task, c), start=1):
            cell = ws.cell(row=row, column=col, value=value)
            _apply_style(cell, styles[col - 1])
        ws.row_dimensions[row].height = row_height


def _row_to_cells(task: WBSTask, c: _Computed) -> list[str]:
    """태스크 1건을 템플릿 열 순서의 값 목록으로 변환한다."""
    pred_numbers = ", ".join(c.number[p] for p in task.predecessors)
    return [
        c.number[task.id],
        task.name,
        task.role,
        c.start[task.id].isoformat(),
        c.end[task.id].isoformat(),
        _fmt_effort(c.effort[task.id]),
        pred_numbers,
        task.deliverable,
    ]


def _capture_style(cell: Cell) -> dict[str, Any]:
    # Any 사용 사유: Font/Border 등 서로 다른 openpyxl 스타일 타입을 속성명으로 묶어 다룬다
    return {attr: copy(getattr(cell, attr)) for attr in _STYLE_ATTRS}


def _apply_style(cell: Cell, style: dict[str, Any]) -> None:
    for attr, value in style.items():
        setattr(cell, attr, value)
