"""테스트시나리오 엑셀 렌더러.

순수 함수: (검증된 TestScenarioDocument, 템플릿 경로) → 출력 .xlsx 파일.
템플릿의 서식 기준 행을 복제해 값만 주입하며, 서식을 코드로 새로 그리지 않는다.
"""

from copy import copy
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet

from app.exceptions import RenderError
from app.schemas.test_scenario import TestCase, TestScenarioDocument

# 템플릿 구조 상수 (backend/templates/test_scenario.xlsx 와 일치해야 함)
STYLE_ROW = 9  # 서식 기준 행 — 이 행의 서식을 데이터 행 전체에 복제한다
NUM_COLUMNS = 10
UNIT_SHEET = "단위테스트"
INTEGRATION_SHEET = "통합테스트"

# 표지 정보 영역의 값 셀 위치
COVER_CELLS = {
    "project_name": "B5",
    "system_name": "G5",
    "author": "B6",
    "written_date": "G6",
}

# 셀 서식 복제 대상 속성 (openpyxl StyleProxy 는 copy 후 대입해야 한다)
_STYLE_ATTRS = ("font", "border", "fill", "alignment", "number_format", "protection")


def render_test_scenario(doc: TestScenarioDocument, template_path: Path, output_path: Path) -> Path:
    """템플릿에 표지 정보와 테스트케이스를 주입해 출력 파일을 생성한다."""
    if not template_path.is_file():
        raise RenderError(f"템플릿 파일이 없습니다: {template_path}")

    wb = load_workbook(template_path)
    sheet_plan = (
        (UNIT_SHEET, doc.unit_test_cases),
        (INTEGRATION_SHEET, doc.integration_test_cases),
    )
    for sheet_name, cases in sheet_plan:
        if sheet_name not in wb.sheetnames:
            raise RenderError(f"템플릿에 '{sheet_name}' 시트가 없습니다: {template_path}")
        ws = wb[sheet_name]
        _fill_cover(ws, doc)
        _fill_cases(ws, cases)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path


def _fill_cover(ws: Worksheet, doc: TestScenarioDocument) -> None:
    """표지 정보 영역(프로젝트명/시스템명/작성자/작성일)에 값을 주입한다."""
    ws[COVER_CELLS["project_name"]] = doc.project_name
    ws[COVER_CELLS["system_name"]] = doc.system_name
    ws[COVER_CELLS["author"]] = doc.author
    ws[COVER_CELLS["written_date"]] = doc.written_date.isoformat()


def _fill_cases(ws: Worksheet, cases: list[TestCase]) -> None:
    """서식 기준 행의 서식을 복제하면서 테스트케이스를 행 단위로 주입한다."""
    styles = [
        _capture_style(ws.cell(row=STYLE_ROW, column=col)) for col in range(1, NUM_COLUMNS + 1)
    ]
    row_height = ws.row_dimensions[STYLE_ROW].height

    for offset, case in enumerate(cases):
        row = STYLE_ROW + offset
        for col, value in enumerate(_case_to_row(case), start=1):
            cell = ws.cell(row=row, column=col, value=value)
            _apply_style(cell, styles[col - 1])
        ws.row_dimensions[row].height = row_height


def _case_to_row(case: TestCase) -> list[str]:
    """테스트케이스 1건을 템플릿 열 순서의 값 목록으로 변환한다."""
    steps = "\n".join(f"{i}. {step}" for i, step in enumerate(case.test_steps, start=1))
    return [
        case.tc_id,
        case.req_id,
        case.category_major,
        case.category_minor,
        case.scenario_name,
        case.precondition,
        steps,
        case.expected_result,
        case.result or "",
        case.note,
    ]


def _capture_style(cell: Cell) -> dict[str, Any]:
    # Any 사용 사유: Font/Border 등 서로 다른 openpyxl 스타일 타입을 속성명으로 묶어 다룬다
    return {attr: copy(getattr(cell, attr)) for attr in _STYLE_ATTRS}


def _apply_style(cell: Cell, style: dict[str, Any]) -> None:
    for attr, value in style.items():
        setattr(cell, attr, value)
