"""인터페이스정의서 엑셀 렌더러.

순수 함수: (검증된 InterfaceSpecDocument, 템플릿 경로) → 출력 .xlsx 파일.
템플릿의 서식 기준 행을 복제해 값만 주입한다 (목록형: 인터페이스별 메시지 항목을 행으로 펼침).
"""

from copy import copy
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet

from app.exceptions import RenderError
from app.schemas.interface_spec import Interface, InterfaceSpecDocument

# 템플릿 구조 상수 (backend/templates/interface_spec.xlsx 와 일치해야 함)
STYLE_ROW = 9
NUM_COLUMNS = 11
SHEET_NAME = "인터페이스정의서"

COVER_CELLS = {
    "project_name": "B5",
    "system_name": "G5",
    "author": "B6",
    "written_date": "G6",
}

_STYLE_ATTRS = ("font", "border", "fill", "alignment", "number_format", "protection")


def render_interface_spec(
    doc: InterfaceSpecDocument, template_path: Path, output_path: Path
) -> Path:
    """템플릿에 표지 정보와 인터페이스·항목 행을 주입해 출력 파일을 생성한다."""
    if not template_path.is_file():
        raise RenderError(f"템플릿 파일이 없습니다: {template_path}")

    wb = load_workbook(template_path)
    if SHEET_NAME not in wb.sheetnames:
        raise RenderError(f"템플릿에 '{SHEET_NAME}' 시트가 없습니다: {template_path}")
    ws = wb[SHEET_NAME]
    _fill_cover(ws, doc)
    _fill_rows(ws, doc)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path


def _fill_cover(ws: Worksheet, doc: InterfaceSpecDocument) -> None:
    ws[COVER_CELLS["project_name"]] = doc.project_name
    ws[COVER_CELLS["system_name"]] = doc.system_name
    ws[COVER_CELLS["author"]] = doc.author
    ws[COVER_CELLS["written_date"]] = doc.written_date.isoformat()


def _fill_rows(ws: Worksheet, doc: InterfaceSpecDocument) -> None:
    """서식 기준 행의 서식을 복제하면서 인터페이스·항목을 행 단위로 주입한다."""
    styles = [
        _capture_style(ws.cell(row=STYLE_ROW, column=col)) for col in range(1, NUM_COLUMNS + 1)
    ]
    row_height = ws.row_dimensions[STYLE_ROW].height

    offset = 0
    for interface in doc.interfaces:
        for column_values in _interface_rows(interface):
            row = STYLE_ROW + offset
            for col, value in enumerate(column_values, start=1):
                cell = ws.cell(row=row, column=col, value=value)
                _apply_style(cell, styles[col - 1])
            ws.row_dimensions[row].height = row_height
            offset += 1


def _interface_rows(interface: Interface) -> list[list[str]]:
    """인터페이스 1개를 메시지 항목 수만큼의 행(템플릿 열 순서)으로 변환한다.

    번호는 인터페이스 내에서 1부터 매기며, 인터페이스 메타 정보는 매 항목 행에 반복 표기한다.
    """
    rows: list[list[str]] = []
    for no, field in enumerate(interface.fields, start=1):
        rows.append(
            [
                str(no),
                interface.interface_id,
                interface.name,
                interface.send_system,
                interface.recv_system,
                interface.method,
                interface.cycle,
                field.name,
                field.data_type,
                "Y" if field.required else "N",
                field.description,
            ]
        )
    return rows


def _capture_style(cell: Cell) -> dict[str, Any]:
    # Any 사용 사유: Font/Border 등 서로 다른 openpyxl 스타일 타입을 속성명으로 묶어 다룬다
    return {attr: copy(getattr(cell, attr)) for attr in _STYLE_ATTRS}


def _apply_style(cell: Cell, style: dict[str, Any]) -> None:
    for attr, value in style.items():
        setattr(cell, attr, value)
