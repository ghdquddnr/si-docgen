"""양식 온보딩 반자동화 — 고객사 엑셀 양식 분석 → 플레이스홀더 위치 제안.

새 고객사 템플릿(.xlsx)을 받으면 렌더러가 값을 주입할 위치(표지 셀·헤더 행·서식 기준 행·
컬럼 매핑)를 휴리스틱으로 추정해 제안한다. 완전 자동이 아니라 **사람 검수용 제안**이며,
LLM·네트워크 없이 결정론적으로 동작한다.

지원 종류(단일 헤더 행 양식): test_scenario / wbs / table_spec / interface_spec.
(RTM 같은 다단 헤더, docx/pptx 는 범위 밖 — 별도 처리.)
"""

from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from app.exceptions import SiDocgenError

# 표지 필드 → 라벨 동의어 (정규화 비교)
COVER_LABELS: dict[str, list[str]] = {
    "project_name": ["프로젝트명", "사업명", "프로젝트"],
    "system_name": ["시스템명", "시스템"],
    "author": ["작성자"],
    "written_date": ["작성일", "작성일자", "일자"],
}

# 종류별 기대 컬럼 헤더 (렌더러 주입 순서). 고객 양식과 라벨이 달라도 느슨하게 매칭한다.
KIND_HEADERS: dict[str, list[str]] = {
    "test_scenario": [
        "TC ID",
        "연관 요건 ID",
        "대분류",
        "중분류",
        "시나리오명",
        "사전조건",
        "테스트 절차",
        "기대 결과",
        "결과",
        "비고",
    ],
    "wbs": [
        "WBS No.",
        "태스크명",
        "담당 역할",
        "시작일",
        "종료일",
        "공수(MD)",
        "선행 태스크",
        "산출물",
    ],
    "table_spec": [
        "테이블 논리명",
        "테이블 물리명",
        "컬럼 논리명",
        "컬럼 물리명",
        "데이터 타입",
        "PK",
        "FK 참조",
        "Null",
        "기본값",
        "설명",
    ],
    "interface_spec": [
        "I/F ID",
        "인터페이스명",
        "송신 시스템",
        "수신 시스템",
        "연계 방식",
        "주기",
        "항목명",
        "데이터 타입",
        "필수",
        "설명",
    ],
}

# 표지/헤더 탐색 범위
_SCAN_ROWS = 25
_SCAN_COLS = 20


class TemplateAnalysisError(SiDocgenError):
    """양식 분석 입력 오류 (파일 없음/미지원 종류 등). CLI 에서 1 로 매핑."""


@dataclass
class ColumnSuggestion:
    """기대 컬럼 1개에 대한 매핑 제안."""

    field_label: str
    column_letter: str | None  # 매칭된 열 (없으면 None)
    matched_text: str | None  # 고객 양식의 실제 헤더 텍스트


@dataclass
class TemplateAnalysis:
    """양식 분석 결과 — 사람 검수용 제안."""

    kind: str
    sheet_name: str
    cover: dict[str, str | None]  # 표지 필드 → 값 셀 좌표(없으면 None)
    header_row: int | None
    style_row: int | None
    columns: list[ColumnSuggestion]
    warnings: list[str] = field(default_factory=list)

    def to_dict(self) -> dict:
        return {
            "kind": self.kind,
            "sheet_name": self.sheet_name,
            "cover": self.cover,
            "header_row": self.header_row,
            "style_row": self.style_row,
            "columns": [
                {"field": c.field_label, "column": c.column_letter, "matched": c.matched_text}
                for c in self.columns
            ],
            "warnings": self.warnings,
        }


def _norm(value: object) -> str:
    """공백 제거 + 소문자 정규화 (한글은 케이스 영향 없음)."""
    return "".join(str(value).split()).lower()


def _loose_match(cell_text: str, label: str) -> bool:
    """느슨한 라벨 매칭: 정규화 후 동일하거나 한쪽이 다른 쪽을 포함."""
    a, b = _norm(cell_text), _norm(label)
    if not a or not b:
        return False
    return a == b or a in b or b in a


def _merged_anchor(ws: Worksheet, row: int, col: int) -> str:
    """(row,col) 이 병합 영역에 속하면 그 좌상단 좌표를, 아니면 자기 좌표를 반환한다."""
    coord = f"{get_column_letter(col)}{row}"
    for rng in ws.merged_cells.ranges:
        if (rng.min_row, rng.min_col) <= (row, col) <= (rng.max_row, rng.max_col):
            return f"{get_column_letter(rng.min_col)}{rng.min_row}"
    return coord


def _find_cover(ws: Worksheet) -> tuple[dict[str, str | None], list[str]]:
    """표지 라벨 셀을 찾아 오른쪽(병합 해소) 값 셀 좌표를 제안한다."""
    cover: dict[str, str | None] = {f: None for f in COVER_LABELS}
    warnings: list[str] = []
    for row in range(1, _SCAN_ROWS + 1):
        for col in range(1, _SCAN_COLS + 1):
            text = ws.cell(row=row, column=col).value
            if not isinstance(text, str) or not text.strip():
                continue
            for field_name, labels in COVER_LABELS.items():
                if cover[field_name] is None and any(_loose_match(text, lbl) for lbl in labels):
                    cover[field_name] = _merged_anchor(ws, row, col + 1)
    for field_name, cell in cover.items():
        if cell is None:
            warnings.append(f"표지 '{field_name}' 라벨을 찾지 못했습니다 (수동 지정 필요)")
    return cover, warnings


def _find_header_row(ws: Worksheet, expected: list[str]) -> tuple[int | None, int]:
    """기대 헤더 라벨이 가장 많이 매칭되는 행을 헤더 행으로 추정한다. (행번호, 매칭수)."""
    best_row, best_score = None, 0
    for row in range(1, _SCAN_ROWS + 1):
        texts = [
            ws.cell(row=row, column=col).value
            for col in range(1, _SCAN_COLS + 1)
            if isinstance(ws.cell(row=row, column=col).value, str)
        ]
        score = sum(any(_loose_match(t, lbl) for t in texts) for lbl in expected)
        if score > best_score:
            best_row, best_score = row, score
    return best_row, best_score


def _map_columns(
    ws: Worksheet, header_row: int, expected: list[str]
) -> tuple[list[ColumnSuggestion], list[str]]:
    """헤더 행에서 기대 컬럼별로 가장 잘 맞는 열을 찾아 매핑한다."""
    header_cells = {
        col: ws.cell(row=header_row, column=col).value
        for col in range(1, _SCAN_COLS + 1)
        if isinstance(ws.cell(row=header_row, column=col).value, str)
    }
    suggestions: list[ColumnSuggestion] = []
    warnings: list[str] = []
    for label in expected:
        match_col = next(
            (col for col, text in header_cells.items() if _loose_match(text, label)), None
        )
        if match_col is None:
            suggestions.append(ColumnSuggestion(label, None, None))
            warnings.append(f"컬럼 '{label}' 에 대응하는 헤더를 찾지 못했습니다")
        else:
            suggestions.append(
                ColumnSuggestion(label, get_column_letter(match_col), header_cells[match_col])
            )
    return suggestions, warnings


def analyze_xlsx_template(path: Path, kind: str, sheet_name: str | None = None) -> TemplateAnalysis:
    """고객 엑셀 양식을 분석해 표지 셀·헤더 행·서식 기준 행·컬럼 매핑을 제안한다."""
    if kind not in KIND_HEADERS:
        raise TemplateAnalysisError(
            f"지원하지 않는 종류: {kind} (지원: {', '.join(sorted(KIND_HEADERS))})"
        )
    if not path.is_file():
        raise TemplateAnalysisError(f"양식 파일이 없습니다: {path}")

    wb = load_workbook(path, data_only=True)
    if sheet_name is not None:
        if sheet_name not in wb.sheetnames:
            raise TemplateAnalysisError(f"시트 '{sheet_name}' 가 없습니다 (시트: {wb.sheetnames})")
        ws = wb[sheet_name]
    else:
        ws = wb.worksheets[0]

    expected = KIND_HEADERS[kind]
    cover, cover_warn = _find_cover(ws)
    header_row, score = _find_header_row(ws, expected)

    warnings = list(cover_warn)
    columns: list[ColumnSuggestion] = []
    style_row: int | None = None
    if header_row is None or score < max(2, len(expected) // 2):
        warnings.append(
            f"헤더 행을 신뢰도 있게 찾지 못했습니다 (최대 매칭 {score}/{len(expected)}). "
            "시트/종류를 확인하세요"
        )
    else:
        style_row = header_row + 1
        columns, col_warn = _map_columns(ws, header_row, expected)
        warnings.extend(col_warn)

    return TemplateAnalysis(
        kind=kind,
        sheet_name=ws.title,
        cover=cover,
        header_row=header_row,
        style_row=style_row,
        columns=columns,
        warnings=warnings,
    )


def format_report(analysis: TemplateAnalysis) -> str:
    """분석 결과를 사람이 읽을 보고서 문자열로 만든다."""
    lines = [
        f"=== 양식 분석: {analysis.kind} (시트 '{analysis.sheet_name}') ===",
        "",
        "[표지 셀 제안]",
    ]
    for field_name, cell in analysis.cover.items():
        lines.append(f"  {field_name:14s} → {cell or '(미발견)'}")
    lines += [
        "",
        f"[헤더 행] {analysis.header_row or '(미발견)'}",
        f"[서식 기준 행] {analysis.style_row or '(미발견)'}",
        "",
        "[컬럼 매핑 제안]",
    ]
    for c in analysis.columns:
        loc = f"{c.column_letter}열" if c.column_letter else "(미발견)"
        matched = f" ← '{c.matched_text}'" if c.matched_text else ""
        lines.append(f"  {c.field_label:14s} → {loc}{matched}")
    if analysis.warnings:
        lines += ["", "[검토 필요]"]
        lines += [f"  ⚠ {w}" for w in analysis.warnings]
    return "\n".join(lines)
