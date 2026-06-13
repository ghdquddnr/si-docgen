"""요건추적표(RTM) 엑셀 양식(backend/templates/rtm.xlsx) 제작 스크립트.

템플릿 파일을 재현 가능하게 만들기 위한 일회성 제작 도구이며,
런타임 렌더러는 이 스크립트가 아니라 생성된 .xlsx 파일을 사용한다.

구조: 단일 시트 '요건추적표'. 데이터 헤더는 2행 구성으로,
앞 5열(요건ID/요건명/화면ID/프로그램ID/TC ID)은 세로 병합,
뒤 4열은 '단계별 반영 여부' 그룹 헤더 아래 분석/설계/구현/시험 하위 헤더를 둔다.

실행: uv run python scripts/templates/build_rtm_template.py
"""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

OUTPUT_PATH = Path(__file__).resolve().parents[2] / "backend" / "templates" / "rtm.xlsx"

FONT_NAME = "맑은 고딕"

HEADER_FILL = PatternFill("solid", start_color="D9D9D9")  # 데이터 헤더 (진회색)
LABEL_FILL = PatternFill("solid", start_color="F2F2F2")  # 표지 라벨 (연회색)

THIN = Side(style="thin", color="000000")
MEDIUM = Side(style="medium", color="000000")
BORDER_ALL_THIN = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# 데이터 헤더 (열 순서 = 렌더러 주입 순서). 앞 5열 + 단계 4열 = 9열
BASE_HEADERS = [
    ("요건 ID", 14),
    ("요건명", 32),
    ("화면 ID", 16),
    ("프로그램 ID", 18),
    ("TC ID", 18),
]
STAGE_HEADERS = [
    ("분석", 8),
    ("설계", 8),
    ("구현", 8),
    ("시험", 8),
]
NUM_BASE = len(BASE_HEADERS)  # 5
NUM_COLUMNS = NUM_BASE + len(STAGE_HEADERS)  # 9

# 앞 5열 중 줄바꿈(좌측 상단) 정렬할 열, 나머지는 가운데 정렬
WRAP_COLUMNS = {2, 3, 4, 5}  # 요건명/화면ID/프로그램ID/TC ID

GROUP_HEADER_ROW = 8  # 그룹 헤더 행 (앞 5열 세로 병합 시작 + '단계별 반영 여부' 가로 병합)
SUB_HEADER_ROW = 9  # 하위 헤더 행 (분석/설계/구현/시험)
STYLE_ROW = 10  # 서식 기준 행 (렌더러가 이 행의 서식을 복제)


def _outline_range(ws: Worksheet, cell_range: str, side: Side = MEDIUM) -> None:
    """병합 영역을 포함한 사각 범위의 바깥 테두리를 지정한 굵기로 두른다."""
    rows = list(ws[cell_range])
    for r_idx, row in enumerate(rows):
        for c_idx, cell in enumerate(row):
            border = Border(
                left=side if c_idx == 0 else cell.border.left,
                right=side if c_idx == len(row) - 1 else cell.border.right,
                top=side if r_idx == 0 else cell.border.top,
                bottom=side if r_idx == len(rows) - 1 else cell.border.bottom,
            )
            cell.border = border


def build_sheet(ws: Worksheet) -> None:
    """시트에 표지 영역·결재란·2행 데이터 헤더·서식 기준 행을 구성한다."""
    # ── 열 너비
    for idx, (_, width) in enumerate(BASE_HEADERS + STAGE_HEADERS, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    # ── 제목 (A1:F3 병합) — 9열이므로 결재란은 G:I
    ws.merge_cells("A1:F3")
    cell = ws["A1"]
    cell.value = "요건추적표 (RTM)"
    cell.font = Font(name=FONT_NAME, size=18, bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 18
    ws.row_dimensions[2].height = 22
    ws.row_dimensions[3].height = 22

    # ── 결재란 (G1:I3): 헤더 행 + 서명 칸(세로 병합)
    for col, label in zip("GHI", ["담당", "검토", "승인"], strict=True):
        head = ws[f"{col}1"]
        head.value = label
        head.font = Font(name=FONT_NAME, size=9, bold=True)
        head.fill = LABEL_FILL
        head.alignment = Alignment(horizontal="center", vertical="center")
        head.border = BORDER_ALL_THIN
        ws.merge_cells(f"{col}2:{col}3")
        for row in (2, 3):
            ws[f"{col}{row}"].border = BORDER_ALL_THIN
    _outline_range(ws, "G1:I3")

    # ── 표지 정보 영역 (5~6행): 라벨 + 값 (test_scenario 와 동일한 셀 배치)
    info_rows = [
        (5, [("A", "프로젝트명", "B:E"), ("F", "시스템명", "G:I")]),
        (6, [("A", "작성자", "B:E"), ("F", "작성일", "G:I")]),
    ]
    for row, pairs in info_rows:
        ws.row_dimensions[row].height = 20
        for label_col, label, value_span in pairs:
            label_cell = ws[f"{label_col}{row}"]
            label_cell.value = label
            label_cell.font = Font(name=FONT_NAME, size=10, bold=True)
            label_cell.fill = LABEL_FILL
            label_cell.alignment = Alignment(horizontal="center", vertical="center")
            start, end = value_span.split(":")
            ws.merge_cells(f"{start}{row}:{end}{row}")
            value_cell = ws[f"{start}{row}"]
            value_cell.font = Font(name=FONT_NAME, size=10)
            value_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            for col_cell in ws[f"{label_col}{row}:{end}{row}"][0]:
                col_cell.border = BORDER_ALL_THIN
    _outline_range(ws, "A5:I6")

    _build_header(ws)
    _build_style_row(ws)

    # ── 보기/인쇄 설정: 헤더 고정, A4 가로, 폭 맞춤
    ws.freeze_panes = f"A{STYLE_ROW}"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_title_rows = f"{GROUP_HEADER_ROW}:{SUB_HEADER_ROW}"


def _build_header(ws: Worksheet) -> None:
    """2행 데이터 헤더: 앞 5열 세로 병합 + 단계 4열 그룹/하위 헤더."""
    ws.row_dimensions[GROUP_HEADER_ROW].height = 22
    ws.row_dimensions[SUB_HEADER_ROW].height = 22

    def style_header(cell) -> None:  # type: ignore[no-untyped-def]
        cell.font = Font(name=FONT_NAME, size=10, bold=True)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER_ALL_THIN

    # 앞 5열: 그룹/하위 헤더 행을 세로 병합하고 라벨을 그룹 헤더 행에 둔다
    for idx, (header, _) in enumerate(BASE_HEADERS, start=1):
        col_letter = get_column_letter(idx)
        ws.merge_cells(f"{col_letter}{GROUP_HEADER_ROW}:{col_letter}{SUB_HEADER_ROW}")
        cell = ws.cell(row=GROUP_HEADER_ROW, column=idx, value=header)
        style_header(cell)
        style_header(ws.cell(row=SUB_HEADER_ROW, column=idx))  # 병합 하단 셀 테두리 유지

    # 단계 4열: 그룹 헤더 '단계별 반영 여부' 가로 병합 + 하위 헤더
    first = NUM_BASE + 1
    last = NUM_COLUMNS
    first_letter = get_column_letter(first)
    last_letter = get_column_letter(last)
    ws.merge_cells(f"{first_letter}{GROUP_HEADER_ROW}:{last_letter}{GROUP_HEADER_ROW}")
    group_cell = ws.cell(row=GROUP_HEADER_ROW, column=first, value="단계별 반영 여부")
    style_header(group_cell)
    for offset in range(1, len(STAGE_HEADERS)):  # 병합된 나머지 셀 테두리 유지
        style_header(ws.cell(row=GROUP_HEADER_ROW, column=first + offset))
    for offset, (header, _) in enumerate(STAGE_HEADERS):
        style_header(ws.cell(row=SUB_HEADER_ROW, column=first + offset, value=header))


def _build_style_row(ws: Worksheet) -> None:
    """서식 기준 행 (값은 비우고 서식만 완성)."""
    ws.row_dimensions[STYLE_ROW].height = 30
    for idx in range(1, NUM_COLUMNS + 1):
        cell = ws.cell(row=STYLE_ROW, column=idx)
        cell.font = Font(name=FONT_NAME, size=10)
        cell.border = BORDER_ALL_THIN
        if idx in WRAP_COLUMNS:
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        else:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def main() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "요건추적표"
    build_sheet(ws)

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_PATH)
    print(f"생성 완료: {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
