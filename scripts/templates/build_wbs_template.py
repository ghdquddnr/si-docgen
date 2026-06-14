"""WBS(작업분해구조) 엑셀 양식(backend/templates/wbs.xlsx) 제작 스크립트.

템플릿을 재현 가능하게 만드는 일회성 도구이며, 런타임 렌더러는 이 스크립트가 아니라
생성된 .xlsx 를 사용한다. 단일 시트 'WBS', 8열 단일 헤더 + 표지/결재란 + 서식 기준 행.

실행: uv run python scripts/templates/build_wbs_template.py
"""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

OUTPUT_PATH = Path(__file__).resolve().parents[2] / "backend" / "templates" / "wbs.xlsx"

FONT_NAME = "맑은 고딕"

HEADER_FILL = PatternFill("solid", start_color="D9D9D9")
LABEL_FILL = PatternFill("solid", start_color="F2F2F2")

THIN = Side(style="thin", color="000000")
MEDIUM = Side(style="medium", color="000000")
BORDER_ALL_THIN = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# 데이터 헤더 (열 순서 = 렌더러 주입 순서). 8열
HEADERS = [
    ("WBS No.", 10),
    ("태스크명", 34),
    ("담당 역할", 14),
    ("시작일", 13),
    ("종료일", 13),
    ("공수(MD)", 10),
    ("선행 태스크", 14),
    ("산출물", 24),
]
NUM_COLUMNS = len(HEADERS)  # 8

# 좌측 정렬(줄바꿈)할 열: 태스크명(2)/산출물(8). 나머지는 가운데 정렬
LEFT_COLUMNS = {2, 8}

HEADER_ROW = 8
STYLE_ROW = 9  # 서식 기준 행 — 렌더러가 이 행 서식을 데이터 행에 복제


def _outline_range(ws: Worksheet, cell_range: str, side: Side = MEDIUM) -> None:
    """병합 영역을 포함한 사각 범위의 바깥 테두리를 지정한 굵기로 두른다."""
    rows = list(ws[cell_range])
    for r_idx, row in enumerate(rows):
        for c_idx, cell in enumerate(row):
            cell.border = Border(
                left=side if c_idx == 0 else cell.border.left,
                right=side if c_idx == len(row) - 1 else cell.border.right,
                top=side if r_idx == 0 else cell.border.top,
                bottom=side if r_idx == len(rows) - 1 else cell.border.bottom,
            )


def build_sheet(ws: Worksheet) -> None:
    """시트에 표지 영역·결재란·헤더·서식 기준 행을 구성한다."""
    for idx, (_, width) in enumerate(HEADERS, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    # ── 제목 (A1:E3 병합), 결재란 F:H
    ws.merge_cells("A1:E3")
    title = ws["A1"]
    title.value = "작업분해구조 (WBS)"
    title.font = Font(name=FONT_NAME, size=18, bold=True)
    title.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 18
    ws.row_dimensions[2].height = 22
    ws.row_dimensions[3].height = 22

    # ── 결재란 (F1:H3): 헤더 행 + 서명 칸(세로 병합)
    for col, label in zip("FGH", ["담당", "검토", "승인"], strict=True):
        head = ws[f"{col}1"]
        head.value = label
        head.font = Font(name=FONT_NAME, size=9, bold=True)
        head.fill = LABEL_FILL
        head.alignment = Alignment(horizontal="center", vertical="center")
        head.border = BORDER_ALL_THIN
        ws.merge_cells(f"{col}2:{col}3")
        for row in (2, 3):
            ws[f"{col}{row}"].border = BORDER_ALL_THIN
    _outline_range(ws, "F1:H3")

    # ── 표지 정보 영역 (5~6행): 라벨 A/E, 값 B:D / F:H
    info_rows = [
        (5, [("A", "프로젝트명", "B:D"), ("E", "시스템명", "F:H")]),
        (6, [("A", "작성자", "B:D"), ("E", "작성일", "F:H")]),
    ]
    for row, pairs in info_rows:
        ws.row_dimensions[row].height = 20
        for label_col, label, value_span in pairs:
            label_cell = ws[f"{label_col}{row}"]
            label_cell.value = label
            label_cell.font = Font(name=FONT_NAME, size=10, bold=True)
            label_cell.fill = LABEL_FILL
            label_cell.alignment = Alignment(horizontal="center", vertical="center")
            label_cell.border = BORDER_ALL_THIN
            start, end = value_span.split(":")
            ws.merge_cells(f"{start}{row}:{end}{row}")
            value_cell = ws[f"{start}{row}"]
            value_cell.font = Font(name=FONT_NAME, size=10)
            value_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            for col_cell in ws[f"{label_col}{row}:{end}{row}"][0]:
                col_cell.border = BORDER_ALL_THIN
    _outline_range(ws, "A5:H6")

    _build_header(ws)
    _build_style_row(ws)

    ws.freeze_panes = f"A{STYLE_ROW}"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_title_rows = f"{HEADER_ROW}:{HEADER_ROW}"


def _build_header(ws: Worksheet) -> None:
    ws.row_dimensions[HEADER_ROW].height = 24
    for idx, (header, _) in enumerate(HEADERS, start=1):
        cell = ws.cell(row=HEADER_ROW, column=idx, value=header)
        cell.font = Font(name=FONT_NAME, size=10, bold=True)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER_ALL_THIN


def _build_style_row(ws: Worksheet) -> None:
    """서식 기준 행 (값은 비우고 서식만 완성)."""
    ws.row_dimensions[STYLE_ROW].height = 22
    for idx in range(1, NUM_COLUMNS + 1):
        cell = ws.cell(row=STYLE_ROW, column=idx)
        cell.font = Font(name=FONT_NAME, size=10)
        cell.border = BORDER_ALL_THIN
        if idx in LEFT_COLUMNS:
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        else:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def main() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "WBS"
    build_sheet(ws)

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_PATH)
    print(f"생성 완료: {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
