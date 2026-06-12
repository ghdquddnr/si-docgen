"""테스트시나리오 엑셀 양식(backend/templates/test_scenario.xlsx) 제작 스크립트.

템플릿 파일을 재현 가능하게 만들기 위한 일회성 제작 도구이며,
런타임 렌더러는 이 스크립트가 아니라 생성된 .xlsx 파일을 사용한다.

실행: uv run python scripts/templates/build_test_scenario_template.py
"""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.worksheet import Worksheet

OUTPUT_PATH = Path(__file__).resolve().parents[2] / "backend" / "templates" / "test_scenario.xlsx"

FONT_NAME = "맑은 고딕"

# 색상
HEADER_FILL = PatternFill("solid", start_color="D9D9D9")  # 데이터 헤더 (진회색)
LABEL_FILL = PatternFill("solid", start_color="F2F2F2")  # 표지 라벨 (연회색)

# 테두리
THIN = Side(style="thin", color="000000")
MEDIUM = Side(style="medium", color="000000")
BORDER_ALL_THIN = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# 데이터 헤더 (열 순서 = 렌더러 주입 순서)
DATA_HEADERS = [
    ("TC ID", 14),
    ("연관 요건 ID", 14),
    ("대분류", 12),
    ("중분류", 12),
    ("시나리오명", 30),
    ("사전조건", 25),
    ("테스트 절차", 42),
    ("기대 결과", 30),
    ("결과", 10),
    ("비고", 14),
]

# 서식 기준 행에서 좌측 상단 정렬(줄바꿈 허용)할 열, 나머지는 가운데 정렬
WRAP_COLUMNS = {5, 6, 7, 8, 10}  # 시나리오명/사전조건/테스트 절차/기대 결과/비고

HEADER_ROW = 8  # 데이터 헤더 행 번호
STYLE_ROW = 9  # 서식 기준 행 번호 (렌더러가 이 행의 서식을 복제)


def _outline_range(ws: Worksheet, cell_range: str, side: Side = MEDIUM) -> None:
    """병합 영역을 포함한 사각 범위의 바깥 테두리를 지정한 굵기로 두른다."""
    rows = list(ws[cell_range])
    for r_idx, row in enumerate(rows):
        for c_idx, cell in enumerate(row):
            border = Border(
                left=side if c_idx == 0 else cell.border.left,
                right=side if c_idx == len(row) - 1 else cell.border.right,
                top=side if r_idx == 0 else cell.border.top,
                bottom=side if r_idx == len(rows) - 1 else cell.border.bottom,
            )
            cell.border = border


def build_sheet(ws: Worksheet, title: str) -> None:
    """시트 1개에 표지 영역·결재란·데이터 헤더·서식 기준 행을 구성한다."""
    # ── 열 너비
    for idx, (_, width) in enumerate(DATA_HEADERS, start=1):
        ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = width

    # ── 제목 (A1:G3 병합)
    ws.merge_cells("A1:G3")
    cell = ws["A1"]
    cell.value = title
    cell.font = Font(name=FONT_NAME, size=18, bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 18
    ws.row_dimensions[2].height = 22
    ws.row_dimensions[3].height = 22

    # ── 결재란 (H1:J3): 헤더 행 + 서명 칸(세로 병합)
    for col, label in zip("HIJ", ["담당", "검토", "승인"], strict=True):
        head = ws[f"{col}1"]
        head.value = label
        head.font = Font(name=FONT_NAME, size=9, bold=True)
        head.fill = LABEL_FILL
        head.alignment = Alignment(horizontal="center", vertical="center")
        head.border = BORDER_ALL_THIN
        ws.merge_cells(f"{col}2:{col}3")
        for row in (2, 3):
            ws[f"{col}{row}"].border = BORDER_ALL_THIN
    _outline_range(ws, "H1:J3")

    # ── 표지 정보 영역 (5~6행): 라벨 + 값
    info_rows = [
        (5, [("A", "프로젝트명", "B:E"), ("F", "시스템명", "G:J")]),
        (6, [("A", "작성자", "B:E"), ("F", "작성일", "G:J")]),
    ]
    for row, pairs in info_rows:
        ws.row_dimensions[row].height = 20
        for label_col, label, value_span in pairs:
            label_cell = ws[f"{label_col}{row}"]
            label_cell.value = label
            label_cell.font = Font(name=FONT_NAME, size=10, bold=True)
            label_cell.fill = LABEL_FILL
            label_cell.alignment = Alignment(horizontal="center", vertical="center")
            start, end = value_span.split(":")
            ws.merge_cells(f"{start}{row}:{end}{row}")
            value_cell = ws[f"{start}{row}"]
            value_cell.font = Font(name=FONT_NAME, size=10)
            value_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            for col_cell in ws[f"{label_col}{row}:{end}{row}"][0]:
                col_cell.border = BORDER_ALL_THIN
    _outline_range(ws, "A5:J6")

    # ── 데이터 헤더 행
    ws.row_dimensions[HEADER_ROW].height = 28
    for idx, (header, _) in enumerate(DATA_HEADERS, start=1):
        cell = ws.cell(row=HEADER_ROW, column=idx, value=header)
        cell.font = Font(name=FONT_NAME, size=10, bold=True)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER_ALL_THIN

    # ── 서식 기준 행 (값은 비우고 서식만 완성)
    ws.row_dimensions[STYLE_ROW].height = 45
    for idx in range(1, len(DATA_HEADERS) + 1):
        cell = ws.cell(row=STYLE_ROW, column=idx)
        cell.font = Font(name=FONT_NAME, size=10)
        cell.border = BORDER_ALL_THIN
        if idx in WRAP_COLUMNS:
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        else:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # ── 보기/인쇄 설정: 헤더 고정, A4 가로, 폭 맞춤
    ws.freeze_panes = f"A{STYLE_ROW}"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_title_rows = f"{HEADER_ROW}:{HEADER_ROW}"


def main() -> None:
    wb = Workbook()
    unit = wb.active
    unit.title = "단위테스트"
    build_sheet(unit, "단위 테스트 시나리오")

    integration = wb.create_sheet("통합테스트")
    build_sheet(integration, "통합 테스트 시나리오")

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_PATH)
    print(f"생성 완료: {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
