"""요구사항정의서를 머리로 둔 4종 체인 테스트 (B1-2).

- build_rtm_from_chain(requirement_spec=...): 정식 요건명 + 요건당 1행(커버리지 공백 포함)
- validate_requirement_consistency: 유령 REQ 참조 거부
- generate_chain(with_requirements=True): docx 포함 4종 + CLI (LLM 모킹)
"""

import json
from pathlib import Path
from typing import Any

import pytest
from docx import Document
from openpyxl import load_workbook

from app import cli
from app.exceptions import ValidationFailedError
from app.pipelines.generate_chain import ChainResult, generate_chain
from app.schemas import test_scenario as ts
from app.schemas.requirement_spec import RequirementSpecDocument
from app.schemas.rtm import build_rtm_from_chain, validate_requirement_consistency
from app.schemas.screen_spec import ScreenSpecDocument

ROOT = Path(__file__).resolve().parents[2]
INPUT = ROOT / "tests" / "fixtures" / "sources" / "sample_requirements.md"

COVER = {"project_name": "P", "system_name": "S", "author": "A", "written_date": "2026-06-14"}


def _case(tc_id: str, req_id: str) -> dict:
    return {
        "tc_id": tc_id,
        "req_id": req_id,
        "category_major": "공통",
        "category_minor": "기능",
        "scenario_name": "시나리오",
        "test_steps": ["단계"],
        "expected_result": "결과",
    }


def _scenario(cases: list[dict]) -> ts.TestScenarioDocument:
    return ts.TestScenarioDocument.model_validate({**COVER, "unit_test_cases": cases})


def _screen(scr_id: str, req_ids: list[str]) -> dict:
    return {
        "screen_id": scr_id,
        "screen_name": "화면",
        "menu_path": "홈 > 화면",
        "req_ids": req_ids,
        "fields": [{"no": 1, "name": "항목", "field_type": "텍스트박스", "required": True}],
    }


def _screen_spec(screens: list[dict]) -> ScreenSpecDocument:
    return ScreenSpecDocument.model_validate({**COVER, "screens": screens})


def _req(req_id: str, name: str) -> dict:
    return {
        "req_id": req_id,
        "name": name,
        "category": "기능",
        "priority": "중",
        "description": "설명",
        "note": "",
    }


def _requirement_spec(requirements: list[dict]) -> RequirementSpecDocument:
    return RequirementSpecDocument.model_validate(
        {**COVER, "doc_no": "REQ-SPEC-2026-001", "requirements": requirements}
    )


# ── 단위: RTM 파생 (요구사항정의서 기준) ───────────────────────────────────


def test_RTM_요건명을_요구사항정의서에서_채움() -> None:
    requirement_spec = _requirement_spec(
        [_req("REQ-001", "사용자 로그인"), _req("REQ-002", "공지사항 관리")]
    )
    scenario = _scenario([_case("TC-001", "REQ-001"), _case("TC-002", "REQ-002")])
    screen_spec = _screen_spec([_screen("SCR-001", ["REQ-001"])])

    rtm = build_rtm_from_chain(scenario, screen_spec, requirement_spec)
    by_req = {row.req_id: row for row in rtm.rows}
    # '대분류 - 중분류' 가 아니라 정식 요건명
    assert by_req["REQ-001"].req_name == "사용자 로그인"
    assert by_req["REQ-002"].req_name == "공지사항 관리"
    assert by_req["REQ-001"].screen_ids == ["SCR-001"]


def test_TC_없는_요건도_커버리지_공백_행으로_노출() -> None:
    # REQ-003 은 요구사항정의서에만 있고 TC/화면이 없음 → 빈 추적 행으로 남아야 한다
    requirement_spec = _requirement_spec(
        [_req("REQ-001", "로그인"), _req("REQ-003", "미구현 요건")]
    )
    scenario = _scenario([_case("TC-001", "REQ-001")])

    rtm = build_rtm_from_chain(scenario, None, requirement_spec)
    by_req = {row.req_id: row for row in rtm.rows}
    assert set(by_req) == {"REQ-001", "REQ-003"}
    assert by_req["REQ-003"].tc_ids == []
    assert by_req["REQ-003"].stage_reflection.test is False
    assert by_req["REQ-001"].stage_reflection.test is True


# ── 단위: 머리 정합성 검증 ─────────────────────────────────────────────────


def test_정합성_부분집합이면_통과() -> None:
    requirement_spec = _requirement_spec([_req("REQ-001", "로그인"), _req("REQ-002", "공지")])
    scenario = _scenario([_case("TC-001", "REQ-001")])
    screen_spec = _screen_spec([_screen("SCR-001", ["REQ-002"])])
    validate_requirement_consistency(requirement_spec, scenario, screen_spec)  # 예외 없음


def test_시나리오의_유령_요건_거부() -> None:
    requirement_spec = _requirement_spec([_req("REQ-001", "로그인")])
    scenario = _scenario([_case("TC-001", "REQ-999")])
    with pytest.raises(ValidationFailedError, match="REQ-999"):
        validate_requirement_consistency(requirement_spec, scenario)


def test_화면의_유령_요건_거부() -> None:
    requirement_spec = _requirement_spec([_req("REQ-001", "로그인")])
    scenario = _scenario([_case("TC-001", "REQ-001")])
    screen_spec = _screen_spec([_screen("SCR-001", ["REQ-888"])])
    with pytest.raises(ValidationFailedError, match="REQ-888"):
        validate_requirement_consistency(requirement_spec, scenario, screen_spec)


# ── e2e: with_requirements 체인 (LLM 모킹) ─────────────────────────────────

MOCK_REQUIREMENT_SPEC: dict[str, Any] = {
    **COVER,
    "doc_no": "REQ-SPEC-2026-001",
    "revisions": [
        {"version": "1.0", "revised_date": "2026-06-14", "author": "A", "description": "최초 작성"}
    ],
    "requirements": [
        _req("REQ-001", "사용자 로그인"),
        _req("REQ-002", "공지사항 관리"),
    ],
}

MOCK_SCENARIO: dict[str, Any] = {
    **COVER,
    "unit_test_cases": [
        {
            "tc_id": "TC-001",
            "req_id": "REQ-001",
            "category_major": "사용자",
            "category_minor": "로그인",
            "scenario_name": "정상 로그인",
            "test_steps": ["접속"],
            "expected_result": "성공",
        }
    ],
    "integration_test_cases": [],
}

MOCK_SCREEN_SPEC: dict[str, Any] = {
    **COVER,
    "screens": [_screen("SCR-001", ["REQ-001"])],
}


@pytest.fixture
def mock_llm(monkeypatch: pytest.MonkeyPatch) -> None:
    def fake(prompt: str, *, system: str | None = None, json_schema=None, model=None) -> str:
        sys_text = system or ""
        if "요구사항정의서" in sys_text:
            payload = MOCK_REQUIREMENT_SPEC
        elif "화면" in sys_text:
            payload = MOCK_SCREEN_SPEC
        else:
            payload = MOCK_SCENARIO
        return json.dumps(payload, ensure_ascii=False)

    monkeypatch.setattr("app.llm.generate.complete_json", fake)


def test_체인_4종_생성_docx_포함(mock_llm: None, tmp_path: Path) -> None:
    result = generate_chain(INPUT, tmp_path, with_requirements=True, **COVER)
    assert isinstance(result, ChainResult)
    assert result.requirement_spec_path is not None
    assert result.requirement_spec_path.is_file()
    assert result.test_scenario_path.is_file()
    assert result.rtm_path.is_file()
    assert result.screen_spec_path.is_file()
    # 요구사항정의서에 2건 → RTM 2행 (TC 없는 REQ-002 도 행으로 노출)
    assert result.requirement_count == 2

    # docx 가 열리고 요건명이 들어갔는지 확인
    text = "\n".join(p.text for p in Document(str(result.requirement_spec_path)).paragraphs)
    assert "사용자 로그인" in text


def test_체인_RTM_정식_요건명_연결(mock_llm: None, tmp_path: Path) -> None:
    result = generate_chain(INPUT, tmp_path, with_requirements=True, **COVER)
    ws = load_workbook(result.rtm_path)["요건추적표"]
    assert ws.cell(row=10, column=1).value == "REQ-001"
    assert ws.cell(row=10, column=2).value == "사용자 로그인"  # 정식 요건명
    assert ws.cell(row=10, column=3).value == "SCR-001"
    # REQ-002 는 TC/화면 없이도 추적 행으로 존재
    assert ws.cell(row=11, column=1).value == "REQ-002"


def test_cli_with_requirements_종료코드_0(mock_llm: None, tmp_path: Path) -> None:
    code = cli.main(
        ["generate", "--input", str(INPUT), "--output", str(tmp_path), "--with-requirements"]
    )
    assert code == 0
    assert (tmp_path / "requirement_spec.docx").is_file()
    assert (tmp_path / "test_scenario.xlsx").is_file()
    assert (tmp_path / "rtm.xlsx").is_file()
    assert (tmp_path / "screen_spec.pptx").is_file()
