"""테이블정의서 LLM 생성 파이프라인 테스트 (LLM 모킹)."""

import json
from pathlib import Path
from typing import Any

import pytest
from openpyxl import load_workbook

from app import cli
from app.pipelines.generate_table_spec import (
    generate_and_render_table_spec,
    generate_table_spec,
)

ROOT = Path(__file__).resolve().parents[2]
INPUT = ROOT / "tests" / "fixtures" / "sources" / "sample_requirements.md"

MOCK_TABLE_SPEC: dict[str, Any] = {
    "project_name": "P",
    "system_name": "S",
    "author": "A",
    "written_date": "2026-06-14",
    "tables": [
        {
            "logical_name": "사용자",
            "physical_name": "TB_USER",
            "description": "사용자 계정",
            "columns": [
                {
                    "logical_name": "사용자 ID",
                    "physical_name": "USER_ID",
                    "data_type": "VARCHAR(20)",
                    "is_pk": True,
                    "is_nullable": False,
                    "description": "PK",
                },
                {
                    "logical_name": "부서 코드",
                    "physical_name": "DEPT_CODE",
                    "data_type": "VARCHAR(10)",
                    "is_pk": False,
                    "is_nullable": True,
                    "fk_ref": "TB_DEPT.DEPT_CODE",
                    "description": "소속",
                },
            ],
        },
        {
            "logical_name": "부서",
            "physical_name": "TB_DEPT",
            "columns": [
                {
                    "logical_name": "부서 코드",
                    "physical_name": "DEPT_CODE",
                    "data_type": "VARCHAR(10)",
                    "is_pk": True,
                    "is_nullable": False,
                    "description": "PK",
                }
            ],
        },
    ],
}

COVER = {"project_name": "P", "system_name": "S", "author": "A", "written_date": "2026-06-14"}


@pytest.fixture
def mock_llm(monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setattr(
        "app.llm.generate.complete_json",
        lambda *a, **k: json.dumps(MOCK_TABLE_SPEC, ensure_ascii=False),
    )


def test_테이블정의서_생성_및_검증(mock_llm: None) -> None:
    doc = generate_table_spec(INPUT, **COVER)
    assert len(doc.tables) == 2
    assert doc.tables[0].physical_name == "TB_USER"
    assert doc.tables[0].columns[1].fk_ref == "TB_DEPT.DEPT_CODE"


def test_진행_콜백_단계_통지(mock_llm: None) -> None:
    stages: list[str] = []
    generate_table_spec(INPUT, **COVER, on_progress=stages.append)
    assert stages == ["parsing", "generating"]


def test_generate_and_render(mock_llm: None, tmp_path: Path) -> None:
    result = generate_and_render_table_spec(INPUT, tmp_path, **COVER)
    assert result.table_spec_path.is_file()
    assert result.table_count == 2
    assert result.column_count == 3  # 2 + 1
    ws = load_workbook(result.table_spec_path)["테이블정의서"]
    assert ws.cell(row=9, column=3).value == "TB_USER"
    assert ws.cell(row=11, column=3).value == "TB_DEPT"  # 9,10=TB_USER, 11=TB_DEPT


def test_cli_table_spec_종료코드_0(mock_llm: None, tmp_path: Path) -> None:
    code = cli.main(["table-spec", "--input", str(INPUT), "--output", str(tmp_path)])
    assert code == 0
    assert (tmp_path / "table_spec.xlsx").is_file()
