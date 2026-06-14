"""웹 체인 잡(with_screens) e2e 테스트 (LLM 모킹).

업로드 → 체인 실행(시나리오 + 화면정의서) → 저장 → 렌더 → pptx 다운로드를 확인한다.
"""

import io
import json
from collections.abc import Iterator
from pathlib import Path
from typing import Any

import pytest
from fastapi.testclient import TestClient
from pptx import Presentation

from app.api.main import app
from app.config import get_settings
from app.db.base import Base
from app.db.models import Job, JobStatus
from app.db.session import SessionLocal, rebind_engine

SCENARIO: dict[str, Any] = {
    "project_name": "P",
    "system_name": "S",
    "author": "A",
    "written_date": "2026-06-14",
    "unit_test_cases": [
        {
            "tc_id": "TC-001",
            "req_id": "REQ-001",
            "category_major": "공통",
            "category_minor": "로그인",
            "scenario_name": "정상",
            "test_steps": ["단계"],
            "expected_result": "결과",
        }
    ],
    "integration_test_cases": [],
}

SCREEN_SPEC: dict[str, Any] = {
    "project_name": "P",
    "system_name": "S",
    "author": "A",
    "written_date": "2026-06-14",
    "screens": [
        {
            "screen_id": "SCR-001",
            "screen_name": "로그인",
            "menu_path": "홈 > 로그인",
            "req_ids": ["REQ-001"],
            "fields": [{"no": 1, "name": "ID", "field_type": "텍스트박스", "required": True}],
        }
    ],
}


@pytest.fixture
def client(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> Iterator[TestClient]:
    monkeypatch.setenv("SIDOCGEN_STORAGE_DIR", str(tmp_path / "jobs"))
    get_settings.cache_clear()
    engine = rebind_engine(f"sqlite:///{tmp_path / 't.db'}")
    Base.metadata.create_all(engine)

    def fake(prompt: str, *, system: str | None = None, json_schema=None, model=None) -> str:
        payload = SCREEN_SPEC if system and "화면" in system else SCENARIO
        return json.dumps(payload, ensure_ascii=False)

    monkeypatch.setattr("app.llm.generate.complete_json", fake)
    yield TestClient(app)
    get_settings.cache_clear()


def _chain_job(client: TestClient) -> str:
    resp = client.post(
        "/jobs",
        files={"file": ("req.md", b"REQ-001", "text/markdown")},
        data={"with_screens": "true"},
    )
    assert resp.status_code == 201
    assert resp.json()["with_screens"] is True
    return resp.json()["id"]


def test_체인_잡_시나리오_화면_저장(client: TestClient) -> None:
    job_id = _chain_job(client)
    got = client.get(f"/jobs/{job_id}")
    assert got.json()["status"] == "succeeded"

    with SessionLocal() as db:
        job = db.get(Job, job_id)
        assert job.status is JobStatus.SUCCEEDED
        assert job.scenario_json is not None
        assert job.screen_spec_json is not None
        assert job.screen_spec_json["screens"][0]["screen_id"] == "SCR-001"


def test_화면정의서_조회(client: TestClient) -> None:
    job_id = _chain_job(client)
    resp = client.get(f"/jobs/{job_id}/screen-spec")
    assert resp.status_code == 200
    assert resp.json()["screens"][0]["req_ids"] == ["REQ-001"]


def test_체인_렌더_후_pptx_다운로드(client: TestClient) -> None:
    job_id = _chain_job(client)
    render = client.post(f"/jobs/{job_id}/render")
    assert render.status_code == 200
    body = render.json()
    assert body["screen_count"] == 1
    assert set(body["downloads"]) == {"test_scenario", "rtm", "screen_spec"}

    pptx = client.get(f"/jobs/{job_id}/download/screen_spec")
    assert pptx.status_code == 200
    assert len(Presentation(io.BytesIO(pptx.content)).slides) >= 1

    # RTM 에 화면 ID 연결 확인
    from openpyxl import load_workbook

    rtm = client.get(f"/jobs/{job_id}/download/rtm")
    ws = load_workbook(io.BytesIO(rtm.content))["요건추적표"]
    assert ws.cell(row=10, column=1).value == "REQ-001"
    assert ws.cell(row=10, column=3).value == "SCR-001"


def test_비체인_잡은_화면_없음_409(client: TestClient) -> None:
    resp = client.post("/jobs", files={"file": ("req.md", b"REQ-001", "text/markdown")})
    job_id = resp.json()["id"]
    assert resp.json()["with_screens"] is False
    assert client.get(f"/jobs/{job_id}/screen-spec").status_code == 409
