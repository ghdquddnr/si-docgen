"""다중 산출물 체인 e2e 테스트 (LLM 모킹).

complete_json 모킹은 system 프롬프트로 시나리오/화면정의서 호출을 구분해 각각 고정 응답을 준다.
"""

import json
from pathlib import Path
from typing import Any

import pytest
from openpyxl import load_workbook
from pptx import Presentation

from app import cli
from app.pipelines.generate_chain import ChainResult, generate_chain

ROOT = Path(__file__).resolve().parents[2]
INPUT = ROOT / "tests" / "fixtures" / "sources" / "sample_requirements.md"

MOCK_SCENARIO: dict[str, Any] = {
    "project_name": "P",
    "system_name": "S",
    "author": "A",
    "written_date": "2026-06-14",
    "unit_test_cases": [
        {
            "tc_id": "TC-001",
            "req_id": "REQ-001",
            "category_major": "사용자 관리",
            "category_minor": "로그인",
            "scenario_name": "정상 로그인",
            "test_steps": ["로그인 화면 접속"],
            "expected_result": "메인 이동",
        },
        {
            "tc_id": "TC-002",
            "req_id": "REQ-002",
            "category_major": "게시판",
            "category_minor": "글쓰기",
            "scenario_name": "게시글 등록",
            "test_steps": ["글쓰기 화면 접속"],
            "expected_result": "목록 표시",
        },
    ],
    "integration_test_cases": [],
}

MOCK_SCREEN_SPEC: dict[str, Any] = {
    "project_name": "P",
    "system_name": "S",
    "author": "A",
    "written_date": "2026-06-14",
    "screens": [
        {
            "screen_id": "SCR-001",
            "screen_name": "로그인",
            "menu_path": "홈 > 로그인",
            "req_ids": ["REQ-001"],
            "fields": [{"no": 1, "name": "ID", "field_type": "텍스트박스", "required": True}],
            "logic": ["ID/PW 검증"],
        },
        {
            "screen_id": "SCR-002",
            "screen_name": "게시글 작성",
            "menu_path": "홈 > 게시판 > 작성",
            "req_ids": ["REQ-002"],
            "fields": [{"no": 1, "name": "제목", "field_type": "텍스트박스", "required": True}],
            "logic": ["등록"],
        },
    ],
}


@pytest.fixture
def mock_llm(monkeypatch: pytest.MonkeyPatch) -> None:
    def fake(prompt: str, *, system: str | None = None, json_schema=None, model=None) -> str:
        payload = MOCK_SCREEN_SPEC if system and "화면" in system else MOCK_SCENARIO
        return json.dumps(payload, ensure_ascii=False)

    monkeypatch.setattr("app.llm.generate.complete_json", fake)


def test_체인_3종_생성(mock_llm: None, tmp_path: Path) -> None:
    result = generate_chain(
        INPUT, tmp_path, project_name="P", system_name="S", author="A", written_date="2026-06-14"
    )
    assert isinstance(result, ChainResult)
    assert result.test_scenario_path.is_file()
    assert result.rtm_path.is_file()
    assert result.screen_spec_path.is_file()
    assert result.screen_count == 2
    assert result.requirement_count == 2

    # pptx 가 열리는지 확인
    assert len(Presentation(str(result.screen_spec_path)).slides) >= 1


def test_RTM_에_화면_ID_연결됨(mock_llm: None, tmp_path: Path) -> None:
    result = generate_chain(
        INPUT, tmp_path, project_name="P", system_name="S", author="A", written_date="2026-06-14"
    )
    ws = load_workbook(result.rtm_path)["요건추적표"]
    # 10행=REQ-001→SCR-001, 11행=REQ-002→SCR-002 (3열=화면 ID)
    assert ws.cell(row=10, column=1).value == "REQ-001"
    assert ws.cell(row=10, column=3).value == "SCR-001"
    assert ws.cell(row=11, column=1).value == "REQ-002"
    assert ws.cell(row=11, column=3).value == "SCR-002"


def test_cli_with_screens_종료코드_0(mock_llm: None, tmp_path: Path) -> None:
    code = cli.main(
        ["generate", "--input", str(INPUT), "--output", str(tmp_path), "--with-screens"]
    )
    assert code == 0
    assert (tmp_path / "test_scenario.xlsx").is_file()
    assert (tmp_path / "rtm.xlsx").is_file()
    assert (tmp_path / "screen_spec.pptx").is_file()
