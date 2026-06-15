"""웹 WBS 잡(with_wbs) e2e 테스트 (LLM 모킹).

업로드(with_wbs) → 시나리오 + WBS 생성·저장 → 렌더 → wbs.xlsx 다운로드를 확인한다.
"""

import io
import json
from collections.abc import Iterator
from pathlib import Path
from typing import Any

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from app.api.main import app
from app.config import get_settings
from app.db.base import Base
from app.db.models import Job, JobStatus
from app.db.session import SessionLocal, rebind_engine

SCENARIO: dict[str, Any] = {
    "project_name": "P",
    "system_name": "S",
    "author": "A",
    "written_date": "2026-06-14",
    "unit_test_cases": [
        {
            "tc_id": "TC-001",
            "req_id": "REQ-001",
            "category_major": "공통",
            "category_minor": "로그인",
            "scenario_name": "정상",
            "test_steps": ["단계"],
            "expected_result": "결과",
        }
    ],
    "integration_test_cases": [],
}

WBS: dict[str, Any] = {
    "project_name": "P",
    "system_name": "S",
    "author": "A",
    "written_date": "2026-06-14",
    "start_date": "2026-07-01",
    "tasks": [
        {
            "id": "analysis",
            "name": "분석",
            "children": [
                {
                    "id": "req-analysis",
                    "name": "요구사항 분석",
                    "role": "PL",
                    "duration_days": 5,
                    "effort_md": 10,
                    "deliverable": "요구사항정의서",
                }
            ],
        }
    ],
}


@pytest.fixture
def client(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> Iterator[TestClient]:
    monkeypatch.setenv("SIDOCGEN_STORAGE_DIR", str(tmp_path / "jobs"))
    get_settings.cache_clear()
    engine = rebind_engine(f"sqlite:///{tmp_path / 't.db'}")
    Base.metadata.create_all(engine)

    def fake(prompt: str, *, system: str | None = None, json_schema=None, model=None) -> str:
        payload = WBS if system and "작업분해구조" in system else SCENARIO
        return json.dumps(payload, ensure_ascii=False)

    monkeypatch.setattr("app.llm.generate.complete_json", fake)
    yield TestClient(app)
    get_settings.cache_clear()


def _wbs_job(client: TestClient) -> str:
    resp = client.post(
        "/jobs",
        files={"file": ("req.md", b"REQ-001", "text/markdown")},
        data={"with_wbs": "true", "start_date": "2026-07-01"},
    )
    assert resp.status_code == 201
    assert resp.json()["with_wbs"] is True
    return resp.json()["id"]


def test_WBS_잡_WBS만_저장(client: TestClient) -> None:
    # 문서별 메뉴: WBS 잡은 시나리오 없이 WBS 만 생성한다
    job_id = _wbs_job(client)
    assert client.get(f"/jobs/{job_id}").json()["status"] == "succeeded"
    with SessionLocal() as db:
        job = db.get(Job, job_id)
        assert job.status is JobStatus.SUCCEEDED
        assert job.scenario_json is None
        assert job.wbs_json is not None
        assert job.wbs_json["tasks"][0]["id"] == "analysis"


def test_WBS_조회(client: TestClient) -> None:
    job_id = _wbs_job(client)
    resp = client.get(f"/jobs/{job_id}/wbs")
    assert resp.status_code == 200
    assert resp.json()["tasks"][0]["name"] == "분석"


def test_렌더_후_wbs_다운로드_계층번호(client: TestClient) -> None:
    job_id = _wbs_job(client)
    render = client.post(f"/jobs/{job_id}/render")
    assert render.status_code == 200
    assert "wbs" in render.json()["downloads"]

    wbs = client.get(f"/jobs/{job_id}/download/wbs")
    assert wbs.status_code == 200
    ws = load_workbook(io.BytesIO(wbs.content))["WBS"]
    # 1행 분석(요약), 2행 1.1 요구사항 분석
    assert ws.cell(row=9, column=1).value == "1"
    assert ws.cell(row=10, column=1).value == "1.1"


def test_비WBS_잡은_WBS_없음_409(client: TestClient) -> None:
    resp = client.post("/jobs", files={"file": ("req.md", b"REQ-001", "text/markdown")})
    job_id = resp.json()["id"]
    assert resp.json()["with_wbs"] is False
    assert client.get(f"/jobs/{job_id}/wbs").status_code == 409
