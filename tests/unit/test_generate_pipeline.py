"""테스트시나리오 + RTM 생성 파이프라인 e2e 테스트 (LLM 모킹).

실제 LLM 은 호출하지 않는다 (CLAUDE.md 테스트 규율). complete_json 을 고정 응답으로
모킹해, 원천 문서 → xlsx 2종 출력 흐름과 RTM 파생·정합성을 검증한다.
"""

import json
from pathlib import Path
from typing import Any

import pytest
from openpyxl import load_workbook

from app import cli
from app.pipelines.generate_test_scenario import generate_test_scenario_and_rtm

ROOT = Path(__file__).resolve().parents[2]
INPUT_DOCX = ROOT / "tests" / "fixtures" / "sources" / "sample_requirements.docx"

# 모킹 LLM 이 돌려줄 고정 시나리오 (표지 값은 파이프라인 인자로 덮어쓰지 않고 그대로 사용됨)
MOCK_SCENARIO: dict[str, Any] = {
    "project_name": "테스트 프로젝트",
    "system_name": "테스트 시스템",
    "author": "작성자",
    "written_date": "2026-06-14",
    "unit_test_cases": [
        {
            "tc_id": "TC-001",
            "req_id": "REQ-001",
            "category_major": "사용자 관리",
            "category_minor": "로그인",
            "scenario_name": "정상 로그인",
            "test_steps": ["로그인 화면 접속", "ID/PW 입력", "로그인 클릭"],
            "expected_result": "메인 화면으로 이동",
        },
        {
            "tc_id": "TC-002",
            "req_id": "REQ-001",
            "category_major": "사용자 관리",
            "category_minor": "로그인",
            "scenario_name": "비밀번호 오류",
            "test_steps": ["로그인 화면 접속", "잘못된 PW 입력", "로그인 클릭"],
            "expected_result": "오류 메시지 표시",
        },
        {
            "tc_id": "TC-003",
            "req_id": "REQ-002",
            "category_major": "게시판",
            "category_minor": "글쓰기",
            "scenario_name": "게시글 등록",
            "test_steps": ["글쓰기 화면 접속", "제목/내용 입력", "등록 클릭"],
            "expected_result": "목록에 새 글 표시",
        },
    ],
    "integration_test_cases": [
        {
            "tc_id": "TC-004",
            "req_id": "REQ-002",
            "category_major": "게시판",
            "category_minor": "통합 흐름",
            "scenario_name": "글 등록 후 알림 발송",
            "test_steps": ["글 등록", "구독자 알림 확인"],
            "expected_result": "구독자에게 알림 발송됨",
        }
    ],
}


@pytest.fixture
def mock_llm(monkeypatch: pytest.MonkeyPatch) -> None:
    """generate_validated 가 사용하는 complete_json 을 고정 JSON 응답으로 대체한다."""

    def fake_complete_json(
        prompt: str, *, system: str | None = None, json_schema=None, model=None
    ) -> str:
        return json.dumps(MOCK_SCENARIO, ensure_ascii=False)

    monkeypatch.setattr("app.llm.generate.complete_json", fake_complete_json)


def test_파이프라인_xlsx_2종_생성(mock_llm: None, tmp_path: Path) -> None:
    result = generate_test_scenario_and_rtm(
        INPUT_DOCX,
        tmp_path,
        project_name="P",
        system_name="S",
        author="A",
        written_date="2026-06-14",
    )

    assert result.test_scenario_path.is_file()
    assert result.rtm_path.is_file()
    assert result.unit_count == 3
    assert result.integration_count == 1
    assert result.requirement_count == 2  # REQ-001, REQ-002


def test_생성된_테스트시나리오_내용(mock_llm: None, tmp_path: Path) -> None:
    result = generate_test_scenario_and_rtm(
        INPUT_DOCX,
        tmp_path,
        project_name="P",
        system_name="S",
        author="A",
        written_date="2026-06-14",
    )
    ws = load_workbook(result.test_scenario_path)["단위테스트"]
    assert ws.cell(row=9, column=1).value == "TC-001"  # 서식 기준 행부터 데이터


def test_파생된_RTM_정합성(mock_llm: None, tmp_path: Path) -> None:
    result = generate_test_scenario_and_rtm(
        INPUT_DOCX,
        tmp_path,
        project_name="P",
        system_name="S",
        author="A",
        written_date="2026-06-14",
    )
    ws = load_workbook(result.rtm_path)["요건추적표"]
    # 첫 추적 행(10행): REQ-001, TC-001/TC-002 매핑
    assert ws.cell(row=10, column=1).value == "REQ-001"
    assert ws.cell(row=10, column=5).value == "TC-001\nTC-002"
    # 분석(6열)·시험(9열) 반영 O, 설계(7열)·구현(8열) 미반영
    assert ws.cell(row=10, column=6).value == "O"
    assert ws.cell(row=10, column=9).value == "O"
    assert ws.cell(row=10, column=7).value in (None, "")


def test_cli_main_종료코드_0(mock_llm: None, tmp_path: Path) -> None:
    code = cli.main(
        [
            "generate",
            "--input",
            str(INPUT_DOCX),
            "--output",
            str(tmp_path),
            "--project-name",
            "P",
            "--author",
            "A",
        ]
    )
    assert code == 0
    assert (tmp_path / "test_scenario.xlsx").is_file()
    assert (tmp_path / "rtm.xlsx").is_file()


def test_cli_미지원_입력_종료코드_1(mock_llm: None, tmp_path: Path) -> None:
    bad = tmp_path / "input.hwp"
    bad.write_text("x", encoding="utf-8")
    code = cli.main(["generate", "--input", str(bad), "--output", str(tmp_path / "out")])
    assert code == 1
