"""양식 보관함(C3) e2e 테스트 — 업로드 검증·폴더·선택 양식 렌더 (LLM 모킹)."""

import io
import json
from collections.abc import Iterator
from pathlib import Path
from typing import Any

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from app.api.main import app
from app.config import get_settings
from app.db.base import Base
from app.db.session import rebind_engine
from app.services import templates_service as svc

WBS: dict[str, Any] = {
    "project_name": "P",
    "system_name": "S",
    "author": "A",
    "written_date": "2026-06-16",
    "start_date": "2026-07-01",
    "tasks": [
        {
            "id": "analysis",
            "name": "분석",
            "children": [
                {
                    "id": "req",
                    "name": "요구사항 분석",
                    "role": "PL",
                    "duration_days": 5,
                    "effort_md": 10,
                    "deliverable": "요건정의서",
                }
            ],
        }
    ],
}


@pytest.fixture
def client(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> Iterator[TestClient]:
    monkeypatch.setenv("SIDOCGEN_STORAGE_DIR", str(tmp_path / "jobs"))
    monkeypatch.setenv("SIDOCGEN_TEMPLATES_DIR", str(tmp_path / "templates"))
    get_settings.cache_clear()
    engine = rebind_engine(f"sqlite:///{tmp_path / 't.db'}")
    Base.metadata.create_all(engine)
    monkeypatch.setattr(
        "app.llm.generate.complete_json",
        lambda *a, **k: json.dumps(WBS, ensure_ascii=False),
    )
    yield TestClient(app)
    get_settings.cache_clear()


def _default_bytes(kind: str) -> bytes:
    return svc.DEFAULT_TEMPLATES[kind].read_bytes()


def test_기본양식_다운로드(client: TestClient) -> None:
    resp = client.get("/templates/default/wbs")
    assert resp.status_code == 200
    assert load_workbook(io.BytesIO(resp.content)).sheetnames  # 유효한 xlsx
    assert client.get("/templates/default/unknown").status_code == 404


def test_라이브러리_종류_목록(client: TestClient) -> None:
    lib = client.get("/templates").json()
    kinds = {k["kind"] for k in lib["kinds"]}
    assert "wbs" in kinds and "requirement_spec" in kinds and "screen_spec" in kinds
    assert lib["templates"] == [] and lib["folders"] == []


def test_양식_업로드_검증통과_후_목록(client: TestClient) -> None:
    # 기본 양식은 구조가 동일하므로 '커스텀'으로 업로드해도 검증 통과
    resp = client.post(
        "/templates",
        files={"file": ("회사_wbs.xlsx", _default_bytes("wbs"), "application/octet-stream")},
        data={"kind": "wbs", "name": "우리회사 WBS"},
    )
    assert resp.status_code == 201
    tid = resp.json()["id"]
    assert resp.json()["kind"] == "wbs"

    lib = client.get("/templates").json()
    assert [t["id"] for t in lib["templates"]] == [tid]


def test_구조_불일치_업로드_400(client: TestClient) -> None:
    # WBS 종류인데 빈/깨진 내용 → 열 수 없어 검증 실패
    resp = client.post(
        "/templates",
        files={"file": ("bad.xlsx", b"not a real xlsx", "application/octet-stream")},
        data={"kind": "wbs"},
    )
    assert resp.status_code == 400


def test_확장자_불일치_400(client: TestClient) -> None:
    # docx 를 wbs(xlsx) 로 업로드 → 확장자 불일치
    resp = client.post(
        "/templates",
        files={"file": ("x.docx", _default_bytes("requirement_spec"), "application/octet-stream")},
        data={"kind": "wbs"},
    )
    assert resp.status_code == 400


def test_폴더_생성_및_하위_양식_삭제(client: TestClient) -> None:
    folder = client.post("/templates/folders", json={"name": "고객사A"})
    assert folder.status_code == 201
    fid = folder.json()["id"]

    up = client.post(
        "/templates",
        files={"file": ("a.xlsx", _default_bytes("table_spec"), "application/octet-stream")},
        data={"kind": "table_spec", "folder_id": fid},
    )
    assert up.status_code == 201
    assert up.json()["folder_id"] == fid

    # 폴더 삭제 시 하위 양식도 함께 제거
    assert client.delete(f"/templates/folders/{fid}").json()["deleted"] is True
    assert client.get("/templates").json()["templates"] == []


def test_없는_폴더에_업로드_404(client: TestClient) -> None:
    resp = client.post(
        "/templates",
        files={"file": ("a.xlsx", _default_bytes("wbs"), "application/octet-stream")},
        data={"kind": "wbs", "folder_id": "nope"},
    )
    assert resp.status_code == 404


def test_선택한_양식으로_렌더(client: TestClient) -> None:
    # 커스텀 WBS 양식 업로드 → 잡 생성 시 선택 → 렌더 성공
    tid = client.post(
        "/templates",
        files={"file": ("wbs.xlsx", _default_bytes("wbs"), "application/octet-stream")},
        data={"kind": "wbs"},
    ).json()["id"]

    job_id = client.post(
        "/jobs",
        files={"file": ("req.md", b"REQ-001", "text/markdown")},
        data={
            "with_wbs": "true",
            "start_date": "2026-07-01",
            "template_ids": json.dumps({"wbs": tid}),
        },
    ).json()["id"]

    render = client.post(f"/jobs/{job_id}/render")
    assert render.status_code == 200
    assert "wbs" in render.json()["downloads"]
    wbs = client.get(f"/jobs/{job_id}/download/wbs")
    assert load_workbook(io.BytesIO(wbs.content)).sheetnames
