"""인터페이스정의서 LLM 생성 파이프라인 테스트 (LLM 모킹)."""

import json
from pathlib import Path
from typing import Any

import pytest
from openpyxl import load_workbook

from app import cli
from app.pipelines.generate_interface_spec import (
    generate_and_render_interface_spec,
    generate_interface_spec,
)

ROOT = Path(__file__).resolve().parents[2]
INPUT = ROOT / "tests" / "fixtures" / "sources" / "sample_requirements.md"

MOCK_INTERFACE_SPEC: dict[str, Any] = {
    "project_name": "P",
    "system_name": "S",
    "author": "A",
    "written_date": "2026-06-14",
    "interfaces": [
        {
            "interface_id": "IF-001",
            "name": "사용자 정보 연계",
            "send_system": "인사시스템",
            "recv_system": "포털",
            "method": "REST API",
            "cycle": "실시간",
            "fields": [
                {
                    "name": "사용자 ID",
                    "data_type": "String(20)",
                    "required": True,
                    "description": "사번",
                },
                {
                    "name": "부서 코드",
                    "data_type": "String(10)",
                    "required": False,
                    "description": "부서",
                },
            ],
        },
        {
            "interface_id": "IF-002",
            "name": "결재 송신",
            "send_system": "포털",
            "recv_system": "회계",
            "method": "MQ",
            "cycle": "일 1회 배치",
            "fields": [
                {
                    "name": "문서 번호",
                    "data_type": "String(30)",
                    "required": True,
                    "description": "문서",
                }
            ],
        },
    ],
}

COVER = {"project_name": "P", "system_name": "S", "author": "A", "written_date": "2026-06-14"}


@pytest.fixture
def mock_llm(monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setattr(
        "app.llm.generate.complete_json",
        lambda *a, **k: json.dumps(MOCK_INTERFACE_SPEC, ensure_ascii=False),
    )


def test_인터페이스정의서_생성_및_검증(mock_llm: None) -> None:
    doc = generate_interface_spec(INPUT, **COVER)
    assert len(doc.interfaces) == 2
    assert doc.interfaces[0].interface_id == "IF-001"
    assert doc.interfaces[1].method == "MQ"


def test_진행_콜백_단계_통지(mock_llm: None) -> None:
    stages: list[str] = []
    generate_interface_spec(INPUT, **COVER, on_progress=stages.append)
    assert stages == ["parsing", "generating"]


def test_generate_and_render(mock_llm: None, tmp_path: Path) -> None:
    result = generate_and_render_interface_spec(INPUT, tmp_path, **COVER)
    assert result.interface_spec_path.is_file()
    assert result.interface_count == 2
    assert result.field_count == 3  # 2 + 1
    ws = load_workbook(result.interface_spec_path)["인터페이스정의서"]
    assert ws.cell(row=9, column=2).value == "IF-001"
    assert ws.cell(row=11, column=2).value == "IF-002"  # 9,10=IF-001, 11=IF-002


def test_cli_interface_spec_종료코드_0(mock_llm: None, tmp_path: Path) -> None:
    code = cli.main(["interface-spec", "--input", str(INPUT), "--output", str(tmp_path)])
    assert code == 0
    assert (tmp_path / "interface_spec.xlsx").is_file()
