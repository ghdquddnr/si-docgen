"""웹 테이블정의서 잡(with_table_spec) e2e 테스트 (LLM 모킹)."""

import io
import json
from collections.abc import Iterator
from pathlib import Path
from typing import Any

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from app.api.main import app
from app.config import get_settings
from app.db.base import Base
from app.db.models import Job, JobStatus
from app.db.session import SessionLocal, rebind_engine

SCENARIO: dict[str, Any] = {
    "project_name": "P",
    "system_name": "S",
    "author": "A",
    "written_date": "2026-06-14",
    "unit_test_cases": [
        {
            "tc_id": "TC-001",
            "req_id": "REQ-001",
            "category_major": "공통",
            "category_minor": "로그인",
            "scenario_name": "정상",
            "test_steps": ["단계"],
            "expected_result": "결과",
        }
    ],
    "integration_test_cases": [],
}

TABLE_SPEC: dict[str, Any] = {
    "project_name": "P",
    "system_name": "S",
    "author": "A",
    "written_date": "2026-06-14",
    "tables": [
        {
            "logical_name": "사용자",
            "physical_name": "TB_USER",
            "columns": [
                {
                    "logical_name": "사용자 ID",
                    "physical_name": "USER_ID",
                    "data_type": "VARCHAR(20)",
                    "is_pk": True,
                    "is_nullable": False,
                    "description": "PK",
                }
            ],
        }
    ],
}


@pytest.fixture
def client(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> Iterator[TestClient]:
    monkeypatch.setenv("SIDOCGEN_STORAGE_DIR", str(tmp_path / "jobs"))
    get_settings.cache_clear()
    engine = rebind_engine(f"sqlite:///{tmp_path / 't.db'}")
    Base.metadata.create_all(engine)

    def fake(prompt: str, *, system: str | None = None, json_schema=None, model=None) -> str:
        payload = TABLE_SPEC if system and "테이블정의서" in system else SCENARIO
        return json.dumps(payload, ensure_ascii=False)

    monkeypatch.setattr("app.llm.generate.complete_json", fake)
    yield TestClient(app)
    get_settings.cache_clear()


def _ts_job(client: TestClient) -> str:
    resp = client.post(
        "/jobs",
        files={"file": ("req.md", b"REQ-001", "text/markdown")},
        data={"with_table_spec": "true"},
    )
    assert resp.status_code == 201
    assert resp.json()["with_table_spec"] is True
    return resp.json()["id"]


def test_테이블정의서_잡_저장(client: TestClient) -> None:
    job_id = _ts_job(client)
    assert client.get(f"/jobs/{job_id}").json()["status"] == "succeeded"
    with SessionLocal() as db:
        job = db.get(Job, job_id)
        assert job.status is JobStatus.SUCCEEDED
        assert job.table_spec_json is not None
        assert job.table_spec_json["tables"][0]["physical_name"] == "TB_USER"


def test_테이블정의서_조회(client: TestClient) -> None:
    job_id = _ts_job(client)
    resp = client.get(f"/jobs/{job_id}/table-spec")
    assert resp.status_code == 200
    assert resp.json()["tables"][0]["logical_name"] == "사용자"


def test_렌더_후_table_spec_다운로드(client: TestClient) -> None:
    job_id = _ts_job(client)
    render = client.post(f"/jobs/{job_id}/render")
    assert render.status_code == 200
    assert "table_spec" in render.json()["downloads"]

    ts = client.get(f"/jobs/{job_id}/download/table_spec")
    assert ts.status_code == 200
    ws = load_workbook(io.BytesIO(ts.content))["테이블정의서"]
    assert ws.cell(row=9, column=3).value == "TB_USER"
    assert ws.cell(row=9, column=7).value == "PK"


def test_비테이블_잡은_409(client: TestClient) -> None:
    resp = client.post("/jobs", files={"file": ("req.md", b"REQ-001", "text/markdown")})
    job_id = resp.json()["id"]
    assert resp.json()["with_table_spec"] is False
    assert client.get(f"/jobs/{job_id}/table-spec").status_code == 409
