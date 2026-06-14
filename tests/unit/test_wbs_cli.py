"""WBS CLI e2e 테스트 (LLM 모킹) — 입력 → wbs.xlsx 출력."""

import json
from pathlib import Path
from typing import Any

import pytest
from openpyxl import load_workbook

from app import cli
from app.pipelines.generate_wbs import generate_and_render_wbs

ROOT = Path(__file__).resolve().parents[2]
INPUT = ROOT / "tests" / "fixtures" / "sources" / "sample_requirements.md"

MOCK_WBS: dict[str, Any] = {
    "project_name": "P",
    "system_name": "S",
    "author": "A",
    "written_date": "2026-06-14",
    "start_date": "2026-07-01",
    "tasks": [
        {
            "id": "analysis",
            "name": "분석",
            "children": [
                {
                    "id": "req-analysis",
                    "name": "요구사항 분석",
                    "role": "PL",
                    "duration_days": 5,
                    "effort_md": 10,
                    "deliverable": "요구사항정의서",
                }
            ],
        },
        {
            "id": "test",
            "name": "통합 시험",
            "role": "QA",
            "duration_days": 5,
            "effort_md": 10,
            "predecessors": ["req-analysis"],
            "deliverable": "테스트결과서",
        },
    ],
}


@pytest.fixture
def mock_llm(monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setattr(
        "app.llm.generate.complete_json",
        lambda *a, **k: json.dumps(MOCK_WBS, ensure_ascii=False),
    )


def test_generate_and_render_wbs(mock_llm: None, tmp_path: Path) -> None:
    result = generate_and_render_wbs(
        INPUT,
        tmp_path,
        project_name="P",
        system_name="S",
        author="A",
        written_date="2026-06-14",
        start_date="2026-07-01",
    )
    assert result.wbs_path.is_file()
    assert result.total_count == 3  # 분석 + 요구사항분석 + 통합시험
    assert result.leaf_count == 2

    ws = load_workbook(result.wbs_path)["WBS"]
    # STYLE_ROW=9: 1행 분석(요약), 2행 1.1 요구사항분석, 3행 2 통합시험
    assert ws.cell(row=9, column=1).value == "1"
    assert ws.cell(row=11, column=1).value == "2"
    # 통합 시험은 선행(req-analysis=1.1) 다음 날 시작
    assert ws.cell(row=11, column=7).value == "1.1"


def test_cli_wbs_종료코드_0(mock_llm: None, tmp_path: Path) -> None:
    code = cli.main(
        [
            "wbs",
            "--input",
            str(INPUT),
            "--output",
            str(tmp_path),
            "--start-date",
            "2026-07-01",
        ]
    )
    assert code == 0
    assert (tmp_path / "wbs.xlsx").is_file()


def test_cli_wbs_미지원_입력_종료코드_1(mock_llm: None, tmp_path: Path) -> None:
    bad = tmp_path / "x.zip"
    bad.write_bytes(b"zip")
    code = cli.main(["wbs", "--input", str(bad), "--output", str(tmp_path)])
    assert code == 1
