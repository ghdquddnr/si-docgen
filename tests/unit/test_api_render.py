"""재렌더링 + 다운로드 엔드포인트 e2e 테스트.

검수(편집) → 재렌더링 → 다운로드 흐름과 렌더링 전 다운로드 거부를 확인한다.
"""

import io
import json
from collections.abc import Iterator
from pathlib import Path
from typing import Any

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from app.api.main import app
from app.config import get_settings
from app.db.base import Base
from app.db.session import rebind_engine

MOCK_SCENARIO: dict[str, Any] = {
    "project_name": "P",
    "system_name": "S",
    "author": "A",
    "written_date": "2026-06-14",
    "unit_test_cases": [
        {
            "tc_id": "TC-001",
            "req_id": "REQ-001",
            "category_major": "공통",
            "category_minor": "로그인",
            "scenario_name": "정상",
            "test_steps": ["단계"],
            "expected_result": "결과",
        }
    ],
    "integration_test_cases": [],
}


@pytest.fixture
def client(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> Iterator[TestClient]:
    monkeypatch.setenv("SIDOCGEN_STORAGE_DIR", str(tmp_path / "jobs"))
    get_settings.cache_clear()
    engine = rebind_engine(f"sqlite:///{tmp_path / 't.db'}")
    Base.metadata.create_all(engine)
    monkeypatch.setattr(
        "app.llm.generate.complete_json",
        lambda *a, **k: json.dumps(MOCK_SCENARIO, ensure_ascii=False),
    )
    yield TestClient(app)
    get_settings.cache_clear()


def _completed_job(client: TestClient) -> str:
    return client.post("/jobs", files={"file": ("req.md", b"REQ-001", "text/markdown")}).json()[
        "id"
    ]


def test_렌더링_후_다운로드(client: TestClient) -> None:
    job_id = _completed_job(client)

    render = client.post(f"/jobs/{job_id}/render")
    assert render.status_code == 200
    body = render.json()
    assert body["unit_count"] == 1
    assert body["requirement_count"] == 1
    assert set(body["downloads"]) == {"test_scenario", "rtm"}

    for kind in ("test_scenario", "rtm"):
        resp = client.get(f"/jobs/{job_id}/download/{kind}")
        assert resp.status_code == 200
        assert resp.headers["content-type"].startswith(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        wb = load_workbook(io.BytesIO(resp.content))  # 유효한 xlsx 인지 확인
        assert wb.sheetnames


def test_렌더링_전_다운로드_409(client: TestClient) -> None:
    job_id = _completed_job(client)
    assert client.get(f"/jobs/{job_id}/download/test_scenario").status_code == 409


def test_알수없는_종류_404(client: TestClient) -> None:
    job_id = _completed_job(client)
    client.post(f"/jobs/{job_id}/render")
    assert client.get(f"/jobs/{job_id}/download/unknown").status_code == 404


def test_편집_후_재렌더링_반영(client: TestClient) -> None:
    job_id = _completed_job(client)
    edited = json.loads(json.dumps(MOCK_SCENARIO))
    edited["unit_test_cases"][0]["scenario_name"] = "수정된 시나리오명"
    assert client.put(f"/jobs/{job_id}/scenario", json=edited).status_code == 200

    client.post(f"/jobs/{job_id}/render")
    resp = client.get(f"/jobs/{job_id}/download/test_scenario")
    wb = load_workbook(io.BytesIO(resp.content))
    ws = wb["단위테스트"]
    # 서식 기준 행(9행) 시나리오명 열(5열)에 편집 내용 반영
    assert ws.cell(row=9, column=5).value == "수정된 시나리오명"


def test_없는_잡_렌더링_404(client: TestClient) -> None:
    assert client.post("/jobs/none/render").status_code == 404
