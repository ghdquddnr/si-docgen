"""WBS 엑셀 렌더러 골든 파일 테스트.

고정 픽스처 JSON 으로 파일을 생성한 뒤 openpyxl 로 셀 값을 추출해 기대값과 비교한다.
특히 렌더러가 계산하는 **계층 번호·일정(선행 반영)·요약 공수 합산**을 기대값으로 못박는다.
"""

import json
from pathlib import Path
from typing import Any

import pytest
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from app.renderers.wbs_renderer import render_wbs
from app.schemas.wbs import WBSDocument

ROOT = Path(__file__).resolve().parents[2]
TEMPLATE_PATH = ROOT / "backend" / "templates" / "wbs.xlsx"
FIXTURE_PATH = Path(__file__).parent / "fixtures" / "wbs_7.json"

SHEET_NAME = "WBS"
STYLE_ROW = 9
NUM_COLUMNS = 8

# 전위 순회 순서의 기대 행 (계층번호/태스크명/담당/시작/종료/공수/선행/산출물).
# 일정은 start_date=2026-07-01 기준 달력일 + 선행 종료 다음 날 시작으로 계산.
EXPECTED_ROWS = [
    ["1", "분석", "", "2026-07-01", "2026-07-13", "18", "", ""],
    ["1.1", "요구사항 분석", "PL", "2026-07-01", "2026-07-07", "10", "", "요구사항정의서"],
    ["1.2", "화면 설계", "기획", "2026-07-08", "2026-07-13", "8", "1.1", "화면정의서"],
    ["2", "개발", "", "2026-07-14", "2026-07-28", "36", "", ""],
    ["2.1", "백엔드 개발", "백엔드 개발자", "2026-07-14", "2026-07-28", "20", "1.2", "API 모듈"],
    [
        "2.2",
        "프론트엔드 개발",
        "프론트 개발자",
        "2026-07-14",
        "2026-07-24",
        "16",
        "1.2",
        "화면 구현체",
    ],
    ["3", "통합 시험", "QA", "2026-07-29", "2026-08-04", "10", "2.1, 2.2", "테스트결과서"],
]

EXPECTED_MERGES = {
    "A1:E3",
    "F2:F3",
    "G2:G3",
    "H2:H3",
    "B5:D5",
    "F5:H5",
    "B6:D6",
    "F6:H6",
}


@pytest.fixture(scope="module")
def fixture_data() -> dict[str, Any]:
    return json.loads(FIXTURE_PATH.read_text(encoding="utf-8"))


@pytest.fixture(scope="module")
def rendered_path(fixture_data: dict[str, Any], tmp_path_factory: pytest.TempPathFactory) -> Path:
    doc = WBSDocument.model_validate(fixture_data)
    out = tmp_path_factory.mktemp("golden") / "wbs_out.xlsx"
    return render_wbs(doc, TEMPLATE_PATH, out)


def read_row(ws: Worksheet, row: int) -> list[str]:
    return [ws.cell(row=row, column=c).value or "" for c in range(1, NUM_COLUMNS + 1)]


def test_표지_정보_값(rendered_path: Path, fixture_data: dict[str, Any]) -> None:
    ws = load_workbook(rendered_path)[SHEET_NAME]
    assert ws["B5"].value == fixture_data["project_name"]
    assert ws["F5"].value == fixture_data["system_name"]
    assert ws["B6"].value == fixture_data["author"]
    assert ws["F6"].value == fixture_data["written_date"]


def test_데이터_행_셀_값_전체_일치(rendered_path: Path) -> None:
    ws = load_workbook(rendered_path)[SHEET_NAME]
    for offset, expected in enumerate(EXPECTED_ROWS):
        excel_row = STYLE_ROW + offset
        assert read_row(ws, excel_row) == expected, f"{excel_row}행 불일치"


def test_행_수_검증(rendered_path: Path) -> None:
    ws = load_workbook(rendered_path)[SHEET_NAME]
    last_row = STYLE_ROW + len(EXPECTED_ROWS) - 1
    assert ws.cell(row=last_row, column=1).value == "3"
    # 데이터 직후 행에는 값이 없어야 한다
    assert all(ws.cell(row=last_row + 1, column=c).value is None for c in range(1, NUM_COLUMNS + 1))


def test_요약_태스크_공수_합산(rendered_path: Path) -> None:
    ws = load_workbook(rendered_path)[SHEET_NAME]
    # 1. 분석 = 1.1(10) + 1.2(8) = 18, 2. 개발 = 2.1(20) + 2.2(16) = 36
    assert ws.cell(row=STYLE_ROW, column=6).value == "18"
    assert ws.cell(row=STYLE_ROW + 3, column=6).value == "36"


def test_선행_반영_일정(rendered_path: Path) -> None:
    ws = load_workbook(rendered_path)[SHEET_NAME]
    # 1.2 화면 설계는 1.1(종료 07-07) 다음 날인 최초 영업일 07-08 시작
    assert ws.cell(row=STYLE_ROW + 2, column=4).value == "2026-07-08"
    # 3. 통합 시험은 선행(2.1 종료 07-28, 2.2 종료 07-24) 중 늦은 07-28 다음 날
    # 최초 영업일인 07-29 시작
    assert ws.cell(row=STYLE_ROW + 6, column=4).value == "2026-07-29"
    assert ws.cell(row=STYLE_ROW + 6, column=7).value == "2.1, 2.2"


def test_공공_공휴일_일정_건너뛰기(tmp_path: Path) -> None:
    # 2026년 개천절: 10/3(토) -> 대체공휴일 10/5(월)
    # 10/2(금) 시작하는 2일짜리 작업 -> 10/2(금) 하루, 10/6(화) 하루 진행되어 10/6 종료되어야 함
    from datetime import date

    from app.renderers.wbs_renderer import render_wbs
    from app.schemas.wbs import WBSDocument, WBSTask

    doc = WBSDocument(
        project_name="개천절 테스트",
        system_name="시스템",
        author="작성자",
        written_date=date(2026, 6, 17),
        start_date=date(2026, 10, 2),
        tasks=[
            WBSTask(
                id="national-holiday-task",
                name="개천절 걸친 작업",
                duration_days=2,
                effort_md=2.0,
            )
        ],
    )
    out = tmp_path / "national_holiday_wbs.xlsx"
    render_wbs(doc, TEMPLATE_PATH, out)

    ws = load_workbook(out)[SHEET_NAME]
    # 시작일: 2026-10-02
    assert ws.cell(row=STYLE_ROW, column=4).value == "2026-10-02"
    # 종료일: 2026-10-06 (10/3~10/5 주말 및 대체공휴일 건너뜀)
    assert ws.cell(row=STYLE_ROW, column=5).value == "2026-10-06"


def test_병합_영역_보존(rendered_path: Path) -> None:
    ws = load_workbook(rendered_path)[SHEET_NAME]
    assert {str(r) for r in ws.merged_cells.ranges} == EXPECTED_MERGES


def test_헤더_훼손_없음(rendered_path: Path) -> None:
    ws = load_workbook(rendered_path)[SHEET_NAME]
    assert ws["A1"].value == "작업분해구조 (WBS)"
    assert [ws[f"{c}1"].value for c in "FGH"] == ["담당", "검토", "승인"]
    assert [ws.cell(row=8, column=c).value for c in range(1, NUM_COLUMNS + 1)] == [
        "WBS No.",
        "태스크명",
        "담당 역할",
        "시작일",
        "종료일",
        "공수(MD)",
        "선행 태스크",
        "산출물",
    ]


def test_서식_기준_행_서식_복제(rendered_path: Path) -> None:
    ws = load_workbook(rendered_path)[SHEET_NAME]
    last_row = STYLE_ROW + len(EXPECTED_ROWS) - 1
    for row in (STYLE_ROW, last_row):
        for col in range(1, NUM_COLUMNS + 1):
            cell = ws.cell(row=row, column=col)
            assert cell.border.left.style == "thin", f"{cell.coordinate} 테두리 누락"
            assert cell.font.name == "맑은 고딕", f"{cell.coordinate} 폰트 불일치"
    # 태스크명(2열)은 좌측 정렬, WBS No.(1열)는 가운데 정렬
    assert ws.cell(row=last_row, column=2).alignment.horizontal == "left"
    assert ws.cell(row=last_row, column=1).alignment.horizontal == "center"
