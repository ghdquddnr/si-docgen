"""인터페이스정의서 엑셀 렌더러 골든 파일 테스트.

고정 픽스처 JSON 으로 파일을 생성한 뒤 openpyxl 로 셀 값을 추출해 기대값과 비교한다.
목록형: 인터페이스별 메시지 항목이 행으로 펼쳐지고, 인터페이스 메타가 항목마다 반복된다.
"""

import json
from pathlib import Path
from typing import Any

import pytest
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from app.renderers.interface_spec_renderer import render_interface_spec
from app.schemas.interface_spec import InterfaceSpecDocument

ROOT = Path(__file__).resolve().parents[2]
TEMPLATE_PATH = ROOT / "backend" / "templates" / "interface_spec.xlsx"
FIXTURE_PATH = Path(__file__).parent / "fixtures" / "interface_spec_2.json"

SHEET_NAME = "인터페이스정의서"
STYLE_ROW = 9
NUM_COLUMNS = 11

# No./IF ID/IF명/송신/수신/연계방식/주기/항목명/타입/필수/설명
EXPECTED_ROWS = [
    [
        "1",
        "IF-001",
        "사용자 정보 연계",
        "인사시스템",
        "통합 업무포털",
        "REST API",
        "실시간",
        "사용자 ID",
        "String(20)",
        "Y",
        "사번",
    ],
    [
        "2",
        "IF-001",
        "사용자 정보 연계",
        "인사시스템",
        "통합 업무포털",
        "REST API",
        "실시간",
        "부서 코드",
        "String(10)",
        "Y",
        "소속 부서",
    ],
    [
        "1",
        "IF-002",
        "결재 결과 송신",
        "통합 업무포털",
        "회계시스템",
        "MQ",
        "일 1회 배치",
        "문서 번호",
        "String(30)",
        "Y",
        "기안 문서 번호",
    ],
]

EXPECTED_MERGES = {
    "A1:H3",
    "I2:I3",
    "J2:J3",
    "K2:K3",
    "B5:E5",
    "G5:K5",
    "B6:E6",
    "G6:K6",
}


@pytest.fixture(scope="module")
def fixture_data() -> dict[str, Any]:
    return json.loads(FIXTURE_PATH.read_text(encoding="utf-8"))


@pytest.fixture(scope="module")
def rendered_path(fixture_data: dict[str, Any], tmp_path_factory: pytest.TempPathFactory) -> Path:
    doc = InterfaceSpecDocument.model_validate(fixture_data)
    out = tmp_path_factory.mktemp("golden") / "interface_spec_out.xlsx"
    return render_interface_spec(doc, TEMPLATE_PATH, out)


def read_row(ws: Worksheet, row: int) -> list[str]:
    return [ws.cell(row=row, column=c).value or "" for c in range(1, NUM_COLUMNS + 1)]


def test_표지_정보_값(rendered_path: Path, fixture_data: dict[str, Any]) -> None:
    ws = load_workbook(rendered_path)[SHEET_NAME]
    assert ws["B5"].value == fixture_data["project_name"]
    assert ws["G5"].value == fixture_data["system_name"]
    assert ws["B6"].value == fixture_data["author"]
    assert ws["G6"].value == fixture_data["written_date"]


def test_데이터_행_셀_값_전체_일치(rendered_path: Path) -> None:
    ws = load_workbook(rendered_path)[SHEET_NAME]
    for offset, expected in enumerate(EXPECTED_ROWS):
        excel_row = STYLE_ROW + offset
        assert read_row(ws, excel_row) == expected, f"{excel_row}행 불일치"


def test_행_수_검증(rendered_path: Path) -> None:
    ws = load_workbook(rendered_path)[SHEET_NAME]
    last_row = STYLE_ROW + len(EXPECTED_ROWS) - 1
    assert ws.cell(row=last_row, column=8).value == "문서 번호"
    assert all(ws.cell(row=last_row + 1, column=c).value is None for c in range(1, NUM_COLUMNS + 1))


def test_인터페이스별_번호_리셋(rendered_path: Path) -> None:
    ws = load_workbook(rendered_path)[SHEET_NAME]
    # IF-001 첫 항목 No.=1, IF-002 첫 항목(11행)도 No.=1 로 리셋
    assert ws.cell(row=9, column=1).value == "1"
    assert ws.cell(row=11, column=1).value == "1"
    assert ws.cell(row=11, column=2).value == "IF-002"


def test_연계방식_주기_필수_표기(rendered_path: Path) -> None:
    ws = load_workbook(rendered_path)[SHEET_NAME]
    assert ws.cell(row=9, column=6).value == "REST API"
    assert ws.cell(row=9, column=7).value == "실시간"
    assert ws.cell(row=9, column=10).value == "Y"
    assert ws.cell(row=11, column=6).value == "MQ"


def test_병합_영역_보존(rendered_path: Path) -> None:
    ws = load_workbook(rendered_path)[SHEET_NAME]
    assert {str(r) for r in ws.merged_cells.ranges} == EXPECTED_MERGES


def test_헤더_훼손_없음(rendered_path: Path) -> None:
    ws = load_workbook(rendered_path)[SHEET_NAME]
    assert ws["A1"].value == "인터페이스정의서"
    assert [ws[f"{c}1"].value for c in "IJK"] == ["담당", "검토", "승인"]
    assert [ws.cell(row=8, column=c).value for c in range(1, NUM_COLUMNS + 1)] == [
        "No.",
        "I/F ID",
        "인터페이스명",
        "송신 시스템",
        "수신 시스템",
        "연계 방식",
        "주기",
        "항목명",
        "데이터 타입",
        "필수",
        "설명",
    ]


def test_서식_기준_행_서식_복제(rendered_path: Path) -> None:
    ws = load_workbook(rendered_path)[SHEET_NAME]
    last_row = STYLE_ROW + len(EXPECTED_ROWS) - 1
    for row in (STYLE_ROW, last_row):
        for col in range(1, NUM_COLUMNS + 1):
            cell = ws.cell(row=row, column=col)
            assert cell.border.left.style == "thin", f"{cell.coordinate} 테두리 누락"
            assert cell.font.name == "맑은 고딕", f"{cell.coordinate} 폰트 불일치"
    # 인터페이스명(3열)은 좌측 정렬, 필수(10열)는 가운데 정렬
    assert ws.cell(row=last_row, column=3).alignment.horizontal == "left"
    assert ws.cell(row=last_row, column=10).alignment.horizontal == "center"
