"""요건추적표(RTM) 엑셀 렌더러 골든 파일 테스트.

고정 픽스처 JSON 으로 파일을 생성한 뒤 openpyxl 로 셀 값을 추출해 기대값과 비교한다.
바이너리 비교는 사용하지 않는다 (내용 비교만 수행).
"""

import json
from pathlib import Path
from typing import Any

import pytest
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from app.renderers.rtm_renderer import render_rtm
from app.schemas.rtm import RTMDocument

ROOT = Path(__file__).resolve().parents[2]
TEMPLATE_PATH = ROOT / "backend" / "templates" / "rtm.xlsx"
FIXTURE_PATH = Path(__file__).parent / "fixtures" / "rtm_10.json"

SHEET_NAME = "요건추적표"
STYLE_ROW = 10
NUM_COLUMNS = 9
STAGE_MARK = "O"

# 템플릿이 보장해야 하는 병합 영역 (제목/결재란/표지/2행 헤더)
EXPECTED_MERGES = {
    "A1:F3",
    "G2:G3",
    "H2:H3",
    "I2:I3",
    "B5:E5",
    "G5:I5",
    "B6:E6",
    "G6:I6",
    "A8:A9",
    "B8:B9",
    "C8:C9",
    "D8:D9",
    "E8:E9",
    "F8:I8",
}


@pytest.fixture(scope="module")
def fixture_data() -> dict[str, Any]:
    # Any 사용 사유: json.loads 결과의 원시 dict 를 그대로 다룬다
    return json.loads(FIXTURE_PATH.read_text(encoding="utf-8"))


@pytest.fixture(scope="module")
def rendered_path(fixture_data: dict[str, Any], tmp_path_factory: pytest.TempPathFactory) -> Path:
    doc = RTMDocument.model_validate(fixture_data)
    out = tmp_path_factory.mktemp("golden") / "rtm_out.xlsx"
    return render_rtm(doc, TEMPLATE_PATH, out)


def expected_row_values(row: dict[str, Any]) -> list[str]:
    """픽스처 JSON 의 추적 행 1건을 엑셀 행 기대값으로 변환한다 (골든 기준 정의)."""
    stage = row["stage_reflection"]
    return [
        row["req_id"],
        row["req_name"],
        "\n".join(row["screen_ids"]),
        "\n".join(row["program_ids"]),
        "\n".join(row["tc_ids"]),
        STAGE_MARK if stage["analysis"] else "",
        STAGE_MARK if stage["design"] else "",
        STAGE_MARK if stage["implementation"] else "",
        STAGE_MARK if stage["test"] else "",
    ]


def read_row(ws: Worksheet, row: int) -> list[str]:
    """셀 값을 문자열 행으로 추출한다 (빈 셀은 빈 문자열)."""
    return [ws.cell(row=row, column=c).value or "" for c in range(1, NUM_COLUMNS + 1)]


def test_표지_정보_값(rendered_path: Path, fixture_data: dict[str, Any]) -> None:
    ws = load_workbook(rendered_path)[SHEET_NAME]
    assert ws["B5"].value == fixture_data["project_name"]
    assert ws["G5"].value == fixture_data["system_name"]
    assert ws["B6"].value == fixture_data["author"]
    assert ws["G6"].value == fixture_data["written_date"]


def test_데이터_행_셀_값_전체_일치(rendered_path: Path, fixture_data: dict[str, Any]) -> None:
    ws = load_workbook(rendered_path)[SHEET_NAME]
    for offset, row in enumerate(fixture_data["rows"]):
        excel_row = STYLE_ROW + offset
        assert read_row(ws, excel_row) == expected_row_values(row), f"{excel_row}행 불일치"


def test_행_수_검증(rendered_path: Path, fixture_data: dict[str, Any]) -> None:
    ws = load_workbook(rendered_path)[SHEET_NAME]
    n = len(fixture_data["rows"])
    last_row = STYLE_ROW + n - 1
    assert ws.cell(row=last_row, column=1).value == fixture_data["rows"][-1]["req_id"]
    # 데이터 직후 행에는 값이 없어야 한다 (초과 주입 방지)
    assert all(ws.cell(row=last_row + 1, column=c).value is None for c in range(1, NUM_COLUMNS + 1))


def test_병합_영역_보존(rendered_path: Path) -> None:
    ws = load_workbook(rendered_path)[SHEET_NAME]
    assert {str(r) for r in ws.merged_cells.ranges} == EXPECTED_MERGES


def test_헤더_훼손_없음(rendered_path: Path) -> None:
    ws = load_workbook(rendered_path)[SHEET_NAME]
    assert ws["A1"].value == "요건추적표 (RTM)"
    assert [ws[f"{c}1"].value for c in "GHI"] == ["담당", "검토", "승인"]
    assert ws["A5"].value == "프로젝트명"
    assert ws["F5"].value == "시스템명"
    assert [ws.cell(row=8, column=c).value for c in range(1, 6)] == [
        "요건 ID",
        "요건명",
        "화면 ID",
        "프로그램 ID",
        "TC ID",
    ]
    assert ws.cell(row=8, column=6).value == "단계별 반영 여부"
    assert [ws.cell(row=9, column=c).value for c in range(6, 10)] == [
        "분석",
        "설계",
        "구현",
        "시험",
    ]


def test_빈_목록_컬럼은_빈_셀(rendered_path: Path, fixture_data: dict[str, Any]) -> None:
    ws = load_workbook(rendered_path)[SHEET_NAME]
    # REQ-040: program_ids/tc_ids 가 빈 목록 → 4·5열이 빈 셀
    for offset, row in enumerate(fixture_data["rows"]):
        if row["req_id"] != "REQ-040":
            continue
        excel_row = STYLE_ROW + offset
        assert ws.cell(row=excel_row, column=4).value in (None, "")
        assert ws.cell(row=excel_row, column=5).value in (None, "")


def test_서식_기준_행_서식_복제(rendered_path: Path, fixture_data: dict[str, Any]) -> None:
    ws = load_workbook(rendered_path)[SHEET_NAME]
    last_row = STYLE_ROW + len(fixture_data["rows"]) - 1
    for row in (STYLE_ROW, last_row):
        for col in range(1, NUM_COLUMNS + 1):
            cell = ws.cell(row=row, column=col)
            assert cell.border.left.style == "thin", f"{cell.coordinate} 테두리 누락"
            assert cell.font.name == "맑은 고딕", f"{cell.coordinate} 폰트 불일치"
    # 요건명(2열)은 좌측 정렬·줄바꿈, 단계 열(6열)은 가운데 정렬
    assert ws.cell(row=last_row, column=2).alignment.horizontal == "left"
    assert ws.cell(row=last_row, column=2).alignment.wrap_text
    assert ws.cell(row=last_row, column=6).alignment.horizontal == "center"
