"""테스트시나리오 엑셀 렌더러 골든 파일 테스트.

고정 픽스처 JSON 으로 파일을 생성한 뒤 openpyxl 로 셀 값을 추출해 기대값과 비교한다.
바이너리 비교는 사용하지 않는다 (내용 비교만 수행).
"""

import json
from pathlib import Path
from typing import Any

import pytest
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

# 클래스명이 Test* 라 pytest 가 테스트로 오인 수집하지 않도록 모듈로 임포트한다
from app.renderers.xlsx_renderer import render_test_scenario
from app.schemas import test_scenario as ts

ROOT = Path(__file__).resolve().parents[2]
TEMPLATE_PATH = ROOT / "backend" / "templates" / "test_scenario.xlsx"
FIXTURE_PATH = Path(__file__).parent / "fixtures" / "test_scenario_30.json"

STYLE_ROW = 9
NUM_COLUMNS = 10

# 템플릿이 보장해야 하는 병합 영역 (제목/표지/결재란)
EXPECTED_MERGES = {"A1:G3", "B5:E5", "B6:E6", "G5:J5", "G6:J6", "H2:H3", "I2:I3", "J2:J3"}


@pytest.fixture(scope="module")
def fixture_data() -> dict[str, Any]:
    # Any 사용 사유: json.loads 결과의 원시 dict 를 그대로 다룬다
    return json.loads(FIXTURE_PATH.read_text(encoding="utf-8"))


@pytest.fixture(scope="module")
def rendered_path(fixture_data: dict[str, Any], tmp_path_factory: pytest.TempPathFactory) -> Path:
    doc = ts.TestScenarioDocument.model_validate(fixture_data)
    out = tmp_path_factory.mktemp("golden") / "test_scenario_out.xlsx"
    return render_test_scenario(doc, TEMPLATE_PATH, out)


def expected_row_values(case: dict[str, Any]) -> list[str]:
    """픽스처 JSON 의 테스트케이스 1건을 엑셀 행 기대값으로 변환한다 (골든 기준 정의)."""
    steps = "\n".join(f"{i}. {step}" for i, step in enumerate(case["test_steps"], start=1))
    return [
        case["tc_id"],
        case["req_id"],
        case["category_major"],
        case["category_minor"],
        case["scenario_name"],
        case["precondition"],
        steps,
        case["expected_result"],
        case["result"] or "",
        case["note"],
    ]


def read_row(ws: Worksheet, row: int) -> list[str]:
    """셀 값을 문자열 행으로 추출한다 (빈 셀은 빈 문자열)."""
    return [ws.cell(row=row, column=c).value or "" for c in range(1, NUM_COLUMNS + 1)]


SHEET_CASES = [("단위테스트", "unit_test_cases"), ("통합테스트", "integration_test_cases")]


@pytest.mark.parametrize(("sheet_name", "cases_key"), SHEET_CASES)
def test_표지_정보_값(
    rendered_path: Path, fixture_data: dict[str, Any], sheet_name: str, cases_key: str
) -> None:
    ws = load_workbook(rendered_path)[sheet_name]
    assert ws["B5"].value == fixture_data["project_name"]
    assert ws["G5"].value == fixture_data["system_name"]
    assert ws["B6"].value == fixture_data["author"]
    assert ws["G6"].value == fixture_data["written_date"]


@pytest.mark.parametrize(("sheet_name", "cases_key"), SHEET_CASES)
def test_데이터_행_셀_값_전체_일치(
    rendered_path: Path, fixture_data: dict[str, Any], sheet_name: str, cases_key: str
) -> None:
    ws = load_workbook(rendered_path)[sheet_name]
    for offset, case in enumerate(fixture_data[cases_key]):
        row = STYLE_ROW + offset
        assert read_row(ws, row) == expected_row_values(case), f"{sheet_name} {row}행 불일치"


@pytest.mark.parametrize(("sheet_name", "cases_key"), SHEET_CASES)
def test_행_수_검증(
    rendered_path: Path, fixture_data: dict[str, Any], sheet_name: str, cases_key: str
) -> None:
    ws = load_workbook(rendered_path)[sheet_name]
    n = len(fixture_data[cases_key])
    last_row = STYLE_ROW + n - 1
    assert ws.cell(row=last_row, column=1).value == fixture_data[cases_key][-1]["tc_id"]
    # 데이터 직후 행에는 값이 없어야 한다 (초과 주입 방지)
    assert all(ws.cell(row=last_row + 1, column=c).value is None for c in range(1, NUM_COLUMNS + 1))


@pytest.mark.parametrize(("sheet_name", "cases_key"), SHEET_CASES)
def test_병합_영역_보존(
    rendered_path: Path, fixture_data: dict[str, Any], sheet_name: str, cases_key: str
) -> None:
    ws = load_workbook(rendered_path)[sheet_name]
    assert {str(r) for r in ws.merged_cells.ranges} == EXPECTED_MERGES


@pytest.mark.parametrize(("sheet_name", "cases_key"), SHEET_CASES)
def test_표지_라벨_훼손_없음(
    rendered_path: Path, fixture_data: dict[str, Any], sheet_name: str, cases_key: str
) -> None:
    ws = load_workbook(rendered_path)[sheet_name]
    assert ws["A5"].value == "프로젝트명"
    assert ws["F5"].value == "시스템명"
    assert ws["A6"].value == "작성자"
    assert ws["F6"].value == "작성일"
    assert [ws[f"{c}1"].value for c in "HIJ"] == ["담당", "검토", "승인"]


@pytest.mark.parametrize(("sheet_name", "cases_key"), SHEET_CASES)
def test_서식_기준_행_서식_복제(
    rendered_path: Path, fixture_data: dict[str, Any], sheet_name: str, cases_key: str
) -> None:
    ws = load_workbook(rendered_path)[sheet_name]
    last_row = STYLE_ROW + len(fixture_data[cases_key]) - 1
    for row in (STYLE_ROW, last_row):
        for col in range(1, NUM_COLUMNS + 1):
            cell = ws.cell(row=row, column=col)
            assert cell.border.left.style == "thin", f"{cell.coordinate} 테두리 누락"
            assert cell.font.name == "맑은 고딕", f"{cell.coordinate} 폰트 불일치"
    # 시나리오명(5열)은 좌측 정렬·줄바꿈, TC ID(1열)는 가운데 정렬
    assert ws.cell(row=last_row, column=5).alignment.horizontal == "left"
    assert ws.cell(row=last_row, column=5).alignment.wrap_text
    assert ws.cell(row=last_row, column=1).alignment.horizontal == "center"
