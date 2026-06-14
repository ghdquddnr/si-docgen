"""테이블정의서 엑셀 렌더러 골든 파일 테스트.

고정 픽스처 JSON 으로 파일을 생성한 뒤 openpyxl 로 셀 값을 추출해 기대값과 비교한다.
목록형: 테이블별 컬럼이 행으로 펼쳐지고, 테이블 논리/물리명이 컬럼마다 반복된다.
"""

import json
from pathlib import Path
from typing import Any

import pytest
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from app.renderers.table_spec_renderer import render_table_spec
from app.schemas.table_spec import TableSpecDocument

ROOT = Path(__file__).resolve().parents[2]
TEMPLATE_PATH = ROOT / "backend" / "templates" / "table_spec.xlsx"
FIXTURE_PATH = Path(__file__).parent / "fixtures" / "table_spec_2.json"

SHEET_NAME = "테이블정의서"
STYLE_ROW = 9
NUM_COLUMNS = 11

# No./테이블논리/테이블물리/컬럼논리/컬럼물리/타입/PK/FK/Null/기본값/설명
EXPECTED_ROWS = [
    [
        "1",
        "사용자",
        "TB_USER",
        "사용자 ID",
        "USER_ID",
        "VARCHAR(20)",
        "PK",
        "",
        "N",
        "",
        "로그인 ID",
    ],
    [
        "2",
        "사용자",
        "TB_USER",
        "비밀번호",
        "PASSWORD",
        "VARCHAR(256)",
        "",
        "",
        "N",
        "",
        "해시 저장",
    ],
    [
        "3",
        "사용자",
        "TB_USER",
        "부서 코드",
        "DEPT_CODE",
        "VARCHAR(10)",
        "",
        "TB_DEPT.DEPT_CODE",
        "Y",
        "",
        "소속 부서",
    ],
    [
        "1",
        "부서",
        "TB_DEPT",
        "부서 코드",
        "DEPT_CODE",
        "VARCHAR(10)",
        "PK",
        "",
        "N",
        "",
        "부서 식별 코드",
    ],
    ["2", "부서", "TB_DEPT", "부서명", "DEPT_NAME", "VARCHAR(100)", "", "", "N", "''", "부서 이름"],
]

EXPECTED_MERGES = {
    "A1:H3",
    "I2:I3",
    "J2:J3",
    "K2:K3",
    "B5:E5",
    "G5:K5",
    "B6:E6",
    "G6:K6",
}


@pytest.fixture(scope="module")
def fixture_data() -> dict[str, Any]:
    return json.loads(FIXTURE_PATH.read_text(encoding="utf-8"))


@pytest.fixture(scope="module")
def rendered_path(fixture_data: dict[str, Any], tmp_path_factory: pytest.TempPathFactory) -> Path:
    doc = TableSpecDocument.model_validate(fixture_data)
    out = tmp_path_factory.mktemp("golden") / "table_spec_out.xlsx"
    return render_table_spec(doc, TEMPLATE_PATH, out)


def read_row(ws: Worksheet, row: int) -> list[str]:
    return [ws.cell(row=row, column=c).value or "" for c in range(1, NUM_COLUMNS + 1)]


def test_표지_정보_값(rendered_path: Path, fixture_data: dict[str, Any]) -> None:
    ws = load_workbook(rendered_path)[SHEET_NAME]
    assert ws["B5"].value == fixture_data["project_name"]
    assert ws["G5"].value == fixture_data["system_name"]
    assert ws["B6"].value == fixture_data["author"]
    assert ws["G6"].value == fixture_data["written_date"]


def test_데이터_행_셀_값_전체_일치(rendered_path: Path) -> None:
    ws = load_workbook(rendered_path)[SHEET_NAME]
    for offset, expected in enumerate(EXPECTED_ROWS):
        excel_row = STYLE_ROW + offset
        assert read_row(ws, excel_row) == expected, f"{excel_row}행 불일치"


def test_행_수_검증(rendered_path: Path) -> None:
    ws = load_workbook(rendered_path)[SHEET_NAME]
    last_row = STYLE_ROW + len(EXPECTED_ROWS) - 1
    assert ws.cell(row=last_row, column=5).value == "DEPT_NAME"
    assert all(ws.cell(row=last_row + 1, column=c).value is None for c in range(1, NUM_COLUMNS + 1))


def test_PK_Null_FK_표기(rendered_path: Path) -> None:
    ws = load_workbook(rendered_path)[SHEET_NAME]
    # 9행: PK=PK, Null=N(불허), FK 없음
    assert ws.cell(row=9, column=7).value == "PK"
    assert ws.cell(row=9, column=9).value == "N"
    assert ws.cell(row=9, column=8).value in (None, "")
    # 11행: FK 참조 표기, Null=Y
    assert ws.cell(row=11, column=8).value == "TB_DEPT.DEPT_CODE"
    assert ws.cell(row=11, column=9).value == "Y"


def test_테이블별_번호_리셋(rendered_path: Path) -> None:
    ws = load_workbook(rendered_path)[SHEET_NAME]
    # TB_USER 첫 컬럼 No.=1, TB_DEPT 첫 컬럼(12행)도 No.=1 로 리셋
    assert ws.cell(row=9, column=1).value == "1"
    assert ws.cell(row=12, column=1).value == "1"
    assert ws.cell(row=12, column=3).value == "TB_DEPT"


def test_병합_영역_보존(rendered_path: Path) -> None:
    ws = load_workbook(rendered_path)[SHEET_NAME]
    assert {str(r) for r in ws.merged_cells.ranges} == EXPECTED_MERGES


def test_헤더_훼손_없음(rendered_path: Path) -> None:
    ws = load_workbook(rendered_path)[SHEET_NAME]
    assert ws["A1"].value == "테이블정의서"
    assert [ws[f"{c}1"].value for c in "IJK"] == ["담당", "검토", "승인"]
    assert [ws.cell(row=8, column=c).value for c in range(1, NUM_COLUMNS + 1)] == [
        "No.",
        "테이블 논리명",
        "테이블 물리명",
        "컬럼 논리명",
        "컬럼 물리명",
        "데이터 타입",
        "PK",
        "FK 참조",
        "Null",
        "기본값",
        "설명",
    ]


def test_서식_기준_행_서식_복제(rendered_path: Path) -> None:
    ws = load_workbook(rendered_path)[SHEET_NAME]
    last_row = STYLE_ROW + len(EXPECTED_ROWS) - 1
    for row in (STYLE_ROW, last_row):
        for col in range(1, NUM_COLUMNS + 1):
            cell = ws.cell(row=row, column=col)
            assert cell.border.left.style == "thin", f"{cell.coordinate} 테두리 누락"
            assert cell.font.name == "맑은 고딕", f"{cell.coordinate} 폰트 불일치"
    # 컬럼 논리명(4열)은 좌측 정렬, PK(7열)는 가운데 정렬
    assert ws.cell(row=last_row, column=4).alignment.horizontal == "left"
    assert ws.cell(row=last_row, column=7).alignment.horizontal == "center"
